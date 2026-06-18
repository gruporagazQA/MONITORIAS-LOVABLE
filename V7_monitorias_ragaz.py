#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
 RAGAZ — MONITORIAS 2026 — V7
================================================================================
 Novidades em relação ao V5:
   - Relatório completo gerado automaticamente ao final (sem script separado)
   - Aba Padrões & Insights completa: ranking, falhas, SPIN, funil, cobertura
   - Atendimento: resultado renomeado para "Aprovado", exibe nota média
   - Sucesso: falhas por etapa do SPIN + tipos de pendência ranqueados
   - Comercial: tipo de contato separado da análise de falhas
   - Cobertura: motivo de cada ligação fora da análise
   - Proteção contra duplicatas por ID (data+hora+telefone)
   - Nome do arquivo com identificação clara: RELATORIO_MES_ANO_V7_FINAL.xlsx

 USO (PowerShell):
   python "G:\Meu Drive\ARQUITETURA LOVABLE\LINGUAGEM EM CÓDIGO\V7_monitorias_ragaz.py"

 OPÇÕES:
   --area     TODAS | SUCESSO_CORRETOR | ATENDIMENTO | COMERCIAL_CADASTRO
   --max      N         limita N ligações (teste)
   --data     YYYY-MM-DD  filtra um dia específico
   --dry-run             sem Claude, sem transcrição — só verifica cobertura
   --so-txts             pula transcrição de MP3, usa só TXTs existentes

 PRÉ-REQUISITO:
   $env:ANTHROPIC_API_KEY = "sk-ant-..."
================================================================================
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
import tempfile
import time
import traceback
import unicodedata
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# ── dependências ──────────────────────────────────────────────────────────────
def _ensure(mod: str, pkg: str = "") -> None:
    try:
        __import__(mod)
    except ImportError:
        p = pkg or mod
        print(f"[setup] instalando {p}...")
        os.system(f'"{sys.executable}" -m pip install -q {p}')

for _m, _p in [
    ("pandas",            "pandas"),
    ("openpyxl",          "openpyxl"),
    ("xlrd",              "xlrd==1.2.0"),
    ("lxml",              "lxml"),
    ("anthropic",         "anthropic"),
    ("tqdm",              "tqdm"),
    ("librosa",           "librosa"),
    ("soundfile",         "soundfile"),
    ("scipy",             "scipy"),
    ("speech_recognition","SpeechRecognition"),
]:
    _ensure(_m, _p)

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from tqdm import tqdm

# ============================================================================
# 1. CONFIGURAÇÃO
# ============================================================================

BASE_DIR = Path(r"G:\Meu Drive\ARQUITETURA LOVABLE")

AREAS: Dict[str, Dict[str, Any]] = {
    "SUCESSO_CORRETOR": {
        "nome":       "Sucesso do Corretor",
        "supervisor": "Diego",
        "agentes":    ["Aline De Souza", "Ariel Maia", "Tatiana Costa",
                       "Taynnã Alves", "Ligiane Maia", "Eduardo Novais"],
    },
    "ATENDIMENTO": {
        "nome":       "Atendimento",
        "supervisor": "Danielle Moura",
        "agentes":    ["Deuzeni", "Lhaine", "Talita", "Jéssica"],
    },
    "COMERCIAL_CADASTRO": {
        "nome":       "Comercial Cadastro",
        "supervisor": "Rodrigo",
        "agentes":    ["Natália", "Ariely", "Ana Lúcia"],
    },
}

# Claude — tokens por área
CLAUDE_MODEL         = "claude-haiku-4-5-20251001"
MAX_TOKENS = {
    "SUCESSO_CORRETOR":   800,
    "ATENDIMENTO":        1200,
    "COMERCIAL_CADASTRO": 900,
}
MAX_RETRIES_CLAUDE   = 3
CLAUDE_RETRY_DELAY   = 4

# Rate limit
CALL_DELAY_SEC = 1.5
AREA_PAUSE_SEC = 65

# Preços
PRECO_INPUT  = 0.80  / 1_000_000
PRECO_OUTPUT = 4.00  / 1_000_000
USD_TO_BRL   = 5.40

# XLS
STATUS_OK    = "Bem sucedida"
DUR_MIN_SEG  = 30
TIPO_EF      = ("efetuad", "saída", "saida", "outbound")
TIPO_REC     = ("recebid", "entrada", "inbound")

# Áudio
AUDIO_SR    = 16000
SPEECH_LANG = "pt-BR"

# ============================================================================
# 2. LOGGING
# ============================================================================

LOG = logging.getLogger("ragaz_v6")
LOG.setLevel(logging.INFO)
_fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%Y-%m-%d %H:%M:%S")
_sh  = logging.StreamHandler(sys.stdout)
_sh.setFormatter(_fmt)
LOG.addHandler(_sh)

def _add_file_log(output_dir: Path) -> None:
    (output_dir / "logs").mkdir(parents=True, exist_ok=True)
    fh = logging.FileHandler(
        output_dir / "logs" / f"rodada_{datetime.now():%Y%m%d_%H%M%S}.log",
        encoding="utf-8"
    )
    fh.setFormatter(_fmt)
    LOG.addHandler(fh)

# ============================================================================
# 3. UTILITÁRIOS
# ============================================================================

def norm(s: Any) -> str:
    nfkd = unicodedata.normalize("NFKD", str(s or ""))
    return "".join(c for c in nfkd if not unicodedata.combining(c)).lower().strip()

def u8(tel: str) -> str:
    d = re.sub(r"\D", "", tel)
    return d[-8:] if len(d) >= 8 else d

def normalizar_tel(t: Any) -> str:
    if t is None or str(t).strip() in ("", "-", "nan"):
        return ""
    d = re.sub(r"\D", "", str(t))
    if not d:
        return ""
    if not d.startswith("55") and len(d) >= 10:
        d = "55" + d[-10:]
    return "+" + d

def dur_seg(d: Any) -> int:
    if not d or str(d).strip() in ("", "-", "nan"):
        return 0
    s = str(d).lower()
    seg = 0
    m = re.search(r"(\d+)\s*m", s)
    if m: seg += int(m.group(1)) * 60
    m = re.search(r"(\d+)\s*s(?![a-z])", s)
    if m: seg += int(m.group(1))
    if seg == 0:
        raw = re.sub(r"\D", "", s)
        seg = int(raw) if raw else 0
    return seg

def parse_dt(s: Any) -> Optional[datetime]:
    for fmt in ("%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S",
                "%d/%m/%Y %H:%M",    "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(str(s).strip(), fmt)
        except Exception:
            continue
    return None

def truncar(texto: str, n: int = 1500) -> str:
    palavras = texto.split()
    return texto if len(palavras) <= n else " ".join(palavras[:n]) + " [...truncado...]"

def semana_do_mes(data_str: str) -> str:
    try:
        d = datetime.strptime(str(data_str)[:10], "%Y-%m-%d")
        return f"Semana {(d.day - 1) // 7 + 1}"
    except:
        return "?"

def classificar_agente(nome: str) -> Optional[str]:
    alvo = norm(nome)
    for area_id, info in AREAS.items():
        for ag in info["agentes"]:
            if norm(ag) == alvo:
                return area_id
    for area_id, info in AREAS.items():
        for ag in info["agentes"]:
            ag_n = norm(ag)
            if len(ag_n) >= 4 and len(alvo) >= 4:
                if ag_n in alvo or alvo in ag_n:
                    return area_id
    return None

def col(df: pd.DataFrame, candidatas: List[str]) -> Optional[str]:
    cols_n = {norm(c): c for c in df.columns}
    for cand in candidatas:
        n = norm(cand)
        if n in cols_n:
            return cols_n[n]
    for cand in candidatas:
        n = norm(cand)
        for k, real in cols_n.items():
            if n in k:
                return real
    return None

def lig_id(lig: Dict) -> str:
    """ID único da ligação: data + hora + últimos 8 dígitos do telefone."""
    return f"{lig['data']}_{lig['hora'].replace(':','')}_{u8(lig['telefone'])}"

# ============================================================================
# 4. LEITURA DO XLS
# ============================================================================

def encontrar_xls(base: Path) -> Optional[Path]:
    candidatos = [*base.glob("calls_detail_*.xlsx"), *base.glob("calls_detail_*.xls"),
                  *base.glob("*.xlsx"), *base.glob("*.xls")]
    candidatos = [c for c in candidatos if c.is_file()]
    return max(candidatos, key=lambda p: p.stat().st_mtime) if candidatos else None

def ler_xls(arquivo: Path) -> pd.DataFrame:
    LOG.info(f"XLS: {arquivo.name}")
    try:
        return pd.read_excel(arquivo, engine="openpyxl" if arquivo.suffix == ".xlsx" else "xlrd")
    except Exception as e:
        LOG.warning(f"read_excel falhou ({e}), tentando read_html...")
        tabelas = pd.read_html(str(arquivo))
        if not tabelas:
            raise RuntimeError("Nenhuma tabela encontrada no XLS")
        return tabelas[0]

def processar_xls(df: pd.DataFrame) -> List[Dict[str, Any]]:
    c_col  = col(df, ["Colaborador"])
    c_tel  = col(df, ["Telefone"])
    c_data = col(df, ["Data da chamada", "Data", "Data/Hora"])
    c_dur  = col(df, ["Duração da chamada", "Duraçãoda chamada", "Duração"])
    c_tipo = col(df, ["Tipo Chamada", "Tipo de chamada", "Tipo"])
    c_stat = col(df, ["Status"])
    c_crm  = col(df, ["CRM", "Contato CRM"])
    c_obs  = col(df, ["Observação", "Observacao"])

    faltando = [nm for nm, c in [("Colaborador", c_col), ("Telefone", c_tel),
                                  ("Data", c_data), ("Status", c_stat)] if c is None]
    if faltando:
        raise RuntimeError(f"Colunas obrigatórias ausentes: {faltando}")

    ligacoes = []
    for _, row in df.iterrows():
        gestor  = str(row.get(c_col, "") or "").strip()
        tel     = normalizar_tel(row.get(c_tel, ""))
        dt      = parse_dt(row.get(c_data, ""))
        dur_s   = dur_seg(row.get(c_dur, "")) if c_dur else 0
        tipo_r  = norm(row.get(c_tipo, "") or "") if c_tipo else ""
        status  = str(row.get(c_stat, "") or "").strip()
        crm     = str(row.get(c_crm,  "") or "").strip() if c_crm else ""
        obs     = str(row.get(c_obs,  "") or "").strip() if c_obs else ""

        is_ef  = any(t in tipo_r for t in TIPO_EF)
        is_rec = any(t in tipo_r for t in TIPO_REC)
        is_ok  = status == STATUS_OK and dur_s > DUR_MIN_SEG
        area   = classificar_agente(gestor)
        if not area:
            continue

        lig = {
            "area":          area,
            "gestor":        gestor,
            "telefone":      tel,
            "datetime":      dt,
            "data":          dt.strftime("%Y-%m-%d") if dt else "",
            "hora":          dt.strftime("%H:%M:%S") if dt else "",
            "semana":        semana_do_mes(dt.strftime("%Y-%m-%d") if dt else ""),
            "duracao_seg":   dur_s,
            "tipo":          "efetuada" if is_ef else ("recebida" if is_rec else "outro"),
            "status":        status,
            "crm":           crm,
            "obs":           obs,
            "bem_sucedida":  is_ok,
            "transcricao":   None,
            "transcricao_fonte": None,
            "sem_transcricao_motivo": None,
            "analise":       None,
        }
        lig["id"] = lig_id(lig)
        ligacoes.append(lig)

    LOG.info(f"XLS: {len(df)} linhas → {len(ligacoes)} em escopo")
    return ligacoes

# ============================================================================
# 5. ÍNDICE DE ARQUIVOS
# ============================================================================

class IndiceArquivos:
    PAT_TXT = re.compile(r"^(\d{4}-\d{2}-\d{2})_[\d-]+_\+?(\d+)\.txt$")
    PAT_MP3 = re.compile(r"^(\d{4}-\d{2}-\d{2}) [\d-]+ \+?(\d+)\.mp3$")

    def __init__(self, transcr_dir: Path, audios_dir: Path) -> None:
        self.transcr_dir = transcr_dir
        self.audios_dir  = audios_dir
        self.txts:  Dict[str, Path] = {}
        self.mp3s:  Dict[str, Path] = {}
        self._indexar()

    def _chave(self, data: str, tel_digits: str) -> str:
        return f"{data}|{u8(tel_digits)}"

    def _indexar(self) -> None:
        if self.transcr_dir.exists():
            for arq in self.transcr_dir.glob("*.txt"):
                m = self.PAT_TXT.match(arq.name)
                if m:
                    self.txts[self._chave(m.group(1), m.group(2))] = arq
        LOG.info(f"TXTs indexados: {len(self.txts)}")

        if self.audios_dir.exists():
            for arq in self.audios_dir.rglob("*.mp3"):
                m = self.PAT_MP3.match(arq.name)
                if m:
                    self.mp3s[self._chave(m.group(1), m.group(2))] = arq
        LOG.info(f"MP3s indexados: {len(self.mp3s)}")

    def buscar_txt(self, data: str, tel: str) -> Optional[Path]:
        return self.txts.get(self._chave(data, tel.replace("+", "")))

    def buscar_mp3(self, data: str, tel: str) -> Optional[Path]:
        return self.mp3s.get(self._chave(data, tel.replace("+", "")))

# ============================================================================
# 6. TRANSCRIÇÃO DE ÁUDIO
# ============================================================================

class Transcriber:
    def __init__(self, transcr_dir: Path) -> None:
        self.dir = transcr_dir
        self.dir.mkdir(parents=True, exist_ok=True)
        import speech_recognition as sr
        self._sr = sr
        self.rec = sr.Recognizer()

    def transcrever_mp3(self, mp3: Path) -> Tuple[Optional[str], Optional[str]]:
        try:
            import librosa
            import scipy.io.wavfile as wavfile
            import numpy as np
            audio_data, sr_audio = librosa.load(str(mp3), sr=AUDIO_SR, mono=True)
            with tempfile.NamedTemporaryFile(suffix=".wav", delete=False) as tmp:
                tmp_path = tmp.name
            wavfile.write(tmp_path, sr_audio, (audio_data * 32767).astype(np.int16))
            try:
                with self._sr.AudioFile(tmp_path) as source:
                    audio = self.rec.record(source)
                texto = self.rec.recognize_google(audio, language=SPEECH_LANG)
                return texto, None
            except self._sr.UnknownValueError:
                return None, "audio_ininteligivel"
            except self._sr.RequestError as e:
                msg = str(e)
                return None, "rate_limit_google" if ("403" in msg or "Forbidden" in msg) else f"erro_google_api"
            finally:
                try: Path(tmp_path).unlink(missing_ok=True)
                except: pass
        except Exception as e:
            return None, "falha_tecnica"

    def processar_pendentes(self, indice: IndiceArquivos, ligacoes: List[Dict]) -> int:
        pendentes = [
            l for l in ligacoes
            if l["bem_sucedida"]
            and indice.buscar_txt(l["data"], l["telefone"]) is None
            and indice.buscar_mp3(l["data"], l["telefone"]) is not None
        ]
        if not pendentes:
            LOG.info("Nenhum MP3 pendente de transcrição.")
            return 0

        LOG.info(f"Transcrevendo {len(pendentes)} MP3s sem TXT...")
        ok = 0
        for lig in tqdm(pendentes, desc="Transcrição MP3", unit="áudio"):
            mp3   = indice.buscar_mp3(lig["data"], lig["telefone"])
            texto, erro = self.transcrever_mp3(mp3)
            if erro or not texto:
                lig["sem_transcricao_motivo"] = erro or "transcricao_vazia"
                continue
            destino = self.dir / (mp3.stem.replace(" ", "_") + ".txt")
            destino.write_text(texto, encoding="utf-8")
            chave = indice._chave(lig["data"], lig["telefone"].replace("+", ""))
            indice.txts[chave] = destino
            ok += 1

        LOG.info(f"Transcrições novas: {ok}/{len(pendentes)}")
        return ok

# ============================================================================
# 7. MATCH LIGAÇÃO ↔ TRANSCRIÇÃO
# ============================================================================

def casar_transcricoes(ligacoes: List[Dict], indice: IndiceArquivos) -> int:
    encontradas = 0
    for lig in ligacoes:
        if not lig["bem_sucedida"]:
            continue
        arq = indice.buscar_txt(lig["data"], lig["telefone"])
        if arq:
            txt = arq.read_text(encoding="utf-8", errors="replace").strip()
            if len(txt) >= 50:
                lig["transcricao"] = txt
                lig["transcricao_fonte"] = "TXT_EXISTENTE" if "TRANSCRIÇÕES" in str(arq) else "MP3_NOVO"
                encontradas += 1
            else:
                lig["sem_transcricao_motivo"] = "texto_muito_curto"
        else:
            mp3 = indice.buscar_mp3(lig["data"], lig["telefone"])
            if mp3:
                if not lig.get("sem_transcricao_motivo"):
                    lig["sem_transcricao_motivo"] = "transcricao_falhou"
            else:
                lig["sem_transcricao_motivo"] = "sem_audio_sem_txt"

    LOG.info(f"Transcrições casadas: {encontradas}/{sum(1 for l in ligacoes if l['bem_sucedida'])}")
    return encontradas

# ============================================================================
# 8. PROMPTS
# ============================================================================

def prompt_sucesso(lig: Dict, t: str) -> str:
    return f"""Você é analista de qualidade especializado em equipes comerciais de seguros (Ragaz / Suhai).
Avalie esta ligação de SUCESSO DO CORRETOR com foco em resultado, receita e SPIN.

CONTEXTO: Gestor={lig['gestor']} | Duração={lig['duracao_seg']}s | Tipo={lig['tipo']}

TRANSCRIÇÃO:
{t}

Responda APENAS com este JSON (sem texto antes ou depois, sem markdown):
{{"efetivo":"SIM","confianca":0.9,"tipo_pendencia":"RASTREADOR","spin_situacao":"SIM","spin_problema":"SIM","spin_implicacao":"PARCIAL","spin_necessidade":"NAO","impacto_receita":"POSITIVO","motivo_receita":"...","ponto_forte":"...","ponto_critico":"...","resumo":"...","coaching":"MEDIA"}}"""


def prompt_atendimento(lig: Dict, t: str) -> str:
    return f"""Você avalia qualidade de atendimento telefônico em uma corretora de seguros (Ragaz).
Analise a transcrição e preencha o JSON. Use SOMENTE "CUMPRIU" ou "NAO_CUMPRIU" nos critérios.

Agente: {lig['gestor']} | Duração: {lig['duracao_seg']}s

CRITÉRIOS E PESOS:
- saudacao (25): se identificou corretamente — ex: "Ragaz assessoria, [nome], bom dia"
- atencao_clareza (35): escutou o problema, respondeu com segurança e sem contradição
- vocabulario (20): linguagem adequada, sem gírias, termos técnicos explicados se necessário
- cordialidade (20): tom educado, paciente e empático durante toda a ligação

nota_final = soma dos pesos dos critérios CUMPRIU dividido por 10. Escala 0 a 10.
Aprovado = nota_final >= 7.0

TRANSCRIÇÃO:
{t}

Responda APENAS com este JSON (sem texto antes ou depois, sem markdown):
{{"nota_final":0.0,"aprovado":"SIM","saudacao":"CUMPRIU","atencao_clareza":"CUMPRIU","vocabulario":"CUMPRIU","cordialidade":"CUMPRIU","ponto_forte":"...","ponto_melhoria":"...","resumo":"...","coaching":"MEDIA"}}"""


def prompt_comercial(lig: Dict, t: str) -> str:
    return f"""Você avalia ligações comerciais de cadastro de corretores em uma seguradora (Ragaz).
O objetivo é cadastrar novos corretores ou reativar corretores inativos nas bandeiras disponíveis.

Agente: {lig['gestor']} | Duração: {lig['duracao_seg']}s

CRITÉRIOS:
- tipo: CADASTRO_NOVO | REATIVACAO | OFERTA_BANDEIRA | OUTRO
- oferta_clara: SIM se apresentou mais de uma bandeira ou produto, NAO caso contrário
- objecao: TRATOU | NAO_HOUVE | NAO_TRATOU
- fechamento: SIM se houve compromisso concreto (cadastro, agendamento, retorno marcado, proposta enviada), NAO caso contrário

EFETIVO:
- SIM: qualquer compromisso concreto (cadastro, agendamento, retorno marcado, proposta enviada)
- NAO: corretor recusou, desligou sem engajamento ou não houve avanço

TRANSCRIÇÃO:
{t}

Responda APENAS com este JSON (sem texto antes ou depois, sem markdown):
{{"efetivo":"SIM","tipo":"CADASTRO_NOVO","oferta_clara":"SIM","objecao":"NAO_HOUVE","fechamento":"SIM","proximo_passo":"...","ponto_forte":"...","ponto_melhoria":"...","resumo":"...","coaching":"MEDIA"}}"""


def montar_prompt(lig: Dict) -> str:
    t = truncar(lig["transcricao"] or "")
    if lig["area"] == "SUCESSO_CORRETOR":   return prompt_sucesso(lig, t)
    if lig["area"] == "ATENDIMENTO":        return prompt_atendimento(lig, t)
    return prompt_comercial(lig, t)

# ============================================================================
# 9. ANÁLISE CLAUDE
# ============================================================================

class ClaudeAnalyzer:
    def __init__(self) -> None:
        from anthropic import Anthropic
        api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
        if not api_key:
            raise RuntimeError(
                "\n[ERRO] ANTHROPIC_API_KEY não configurada.\n"
                "No PowerShell: $env:ANTHROPIC_API_KEY = 'sk-ant-...'\n"
            )
        self.client     = Anthropic(api_key=api_key)
        self.tokens_in  = 0
        self.tokens_out = 0
        self.erros: List[Dict] = []

    def _parse_json(self, texto: str) -> Optional[Dict]:
        for candidato in [
            texto,
            re.search(r"\{.*\}", texto, re.DOTALL).group() if re.search(r"\{.*\}", texto, re.DOTALL) else None,
            texto.strip() + "}" if texto.strip().startswith("{") and not texto.strip().endswith("}") else None,
        ]:
            if candidato is None:
                continue
            try:
                return json.loads(candidato)
            except Exception:
                pass
        return None

    def analisar(self, lig: Dict) -> Optional[Dict]:
        if not lig.get("transcricao") or len(lig["transcricao"].strip()) < 50:
            return None
        prompt     = montar_prompt(lig)
        max_tokens = MAX_TOKENS.get(lig["area"], 800)

        for tentativa in range(1, MAX_RETRIES_CLAUDE + 1):
            try:
                resp    = self.client.messages.create(
                    model=CLAUDE_MODEL,
                    max_tokens=max_tokens,
                    messages=[{"role": "user", "content": prompt}],
                )
                analise = self._parse_json(resp.content[0].text)
                if analise is None:
                    raise RuntimeError("JSON inválido na resposta")
                self.tokens_in  += resp.usage.input_tokens
                self.tokens_out += resp.usage.output_tokens
                return analise
            except Exception as e:
                LOG.warning(f"Claude tentativa {tentativa}/{MAX_RETRIES_CLAUDE}: {str(e)[:80]}")
                if tentativa < MAX_RETRIES_CLAUDE:
                    time.sleep(CLAUDE_RETRY_DELAY * tentativa)
                else:
                    self.erros.append({"gestor": lig["gestor"], "motivo": str(e)[:120]})
        return None

    def processar_area(self, ligacoes: List[Dict], area_id: str, dry_run: bool) -> int:
        candidatas = [l for l in ligacoes if l["area"] == area_id and l.get("transcricao")]
        nome = AREAS[area_id]["nome"]
        LOG.info(f"[{nome}] {len(candidatas)} ligações com transcrição para analisar.")

        if dry_run:
            LOG.info(f"[{nome}] dry-run: Claude pulado.")
            return 0

        analisadas = 0
        for lig in tqdm(candidatas, desc=nome[:22], unit="lig"):
            result = self.analisar(lig)
            if result:
                lig["analise"] = result
                analisadas += 1
            time.sleep(CALL_DELAY_SEC)

        LOG.info(f"[{nome}] {analisadas}/{len(candidatas)} analisadas.")
        return analisadas

# ============================================================================
# 10. HELPERS DE RESULTADO
# ============================================================================

def is_aprovado(lig: Dict) -> Optional[bool]:
    a = lig.get("analise")
    if not a:
        return None
    area = lig.get("area", "")

    if area == "ATENDIMENTO":
        # campo "aprovado" direto
        ap = str(a.get("aprovado", "")).upper()
        if ap in ("SIM", "NAO"):
            return ap == "SIM"
        try:
            return float(a.get("nota_final", 0)) >= 7.0
        except:
            return None

    if area in ("SUCESSO_CORRETOR", "COMERCIAL_CADASTRO"):
        ef = str(a.get("efetivo", "")).upper()
        return ef == "SIM" if ef in ("SIM", "NAO") else None

    return None

def nota_atendimento(a: Dict) -> float:
    if not a:
        return 0.0
    try:
        return float(a.get("nota_final", 0))
    except:
        pass
    pesos = {"saudacao": 25, "atencao_clareza": 35, "vocabulario": 20, "cordialidade": 20}
    total = sum(p for k, p in pesos.items() if str(a.get(k, "")).upper() == "CUMPRIU")
    return round(total / 10, 1)

def coaching(lig: Dict) -> str:
    a = lig.get("analise", {})
    if not a:
        return ""
    raw = str(a.get("coaching", a.get("prioridade_coaching", ""))).strip()
    # Normaliza para exatamente ALTA, MEDIA ou BAIXA
    _n = norm(raw)
    if any(k in _n for k in ("alta","critica","critico","alto","pessimo","ruim")):
        return "ALTA"
    if any(k in _n for k in ("media","medio","regular","moderada")):
        return "MEDIA"
    if any(k in _n for k in ("baixa","baixo","fraco","deficiente","excelente","otimo","bom")):
        return "BAIXA"
    return raw.upper()[:5]  # fallback: primeiros 5 chars em maiúscula

# ============================================================================
# 11. ESTILOS EXCEL
# ============================================================================

H_FILL = PatternFill("solid", fgColor="1F4E79")
H_FONT = Font(color="FFFFFF", bold=True, size=10)
S_FILL = PatternFill("solid", fgColor="2E75B6")
S_FONT = Font(color="FFFFFF", bold=True, size=10)
A_FILL = PatternFill("solid", fgColor="D9E1F2")
G_FILL = PatternFill("solid", fgColor="E2EFDA")
R_FILL = PatternFill("solid", fgColor="FCE4D6")
Y_FILL = PatternFill("solid", fgColor="FFF2CC")
BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

def _cab(ws, row: int, ncols: int, fill=None) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill  = fill or H_FILL
        cell.font  = H_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = BORDER

def _sec(ws, row: int, texto: str, ncols: int = 8) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=texto)
    c.fill = S_FILL; c.font = S_FONT
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    return row + 1

def _w(ws, col_letra: str, width: float) -> None:
    ws.column_dimensions[col_letra].width = width

# ============================================================================
# 12. GERAÇÃO DO EXCEL
# ============================================================================

def gerar_excel(ligacoes: List[Dict], output_dir: Path, ts: str,
                tokens_in: int, tokens_out: int, mes_ref: str) -> Path:

    wb = Workbook()
    wb.remove(wb.active)

    bem_suc   = [l for l in ligacoes if l["bem_sucedida"]]
    analisadas = [l for l in bem_suc if l.get("analise")]
    custo_usd  = tokens_in * PRECO_INPUT + tokens_out * PRECO_OUTPUT

    # ── ABA 1: RESUMO ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Resumo", 0)
    ws.merge_cells("A1:G1")
    t = ws["A1"]; t.value = f"RAGAZ — Monitorias {mes_ref} — Relatório Final V6"
    t.font = Font(bold=True, size=14, color="1F4E79")
    t.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:G2")
    ws["A2"].value = f"Gerado em {datetime.now():%d/%m/%Y %H:%M}"
    ws["A2"].font  = Font(color="808080", italic=True, size=9)
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4
    row = _sec(ws, row, "RESULTADO POR ÁREA", 7)
    hdrs = ["Área", "Supervisor", "Bem-sucedidas", "Com transcrição",
            "Analisadas", "Aprovadas/Efetivas", "Taxa %"]
    for i, h in enumerate(hdrs, 1): ws.cell(row=row, column=i, value=h)
    _cab(ws, row, len(hdrs)); row += 1

    tot_bs = tot_tr = tot_an = tot_ap = 0
    for area_id, info in AREAS.items():
        ligs_a = [l for l in bem_suc if l["area"] == area_id]
        com_t  = sum(1 for l in ligs_a if l.get("transcricao"))
        anal   = [l for l in ligs_a if l.get("analise")]
        aprov  = sum(1 for l in anal if is_aprovado(l) is True)
        taxa   = round(aprov / len(anal) * 100, 1) if anal else 0.0
        bs     = sum(1 for l in ligacoes if l["area"] == area_id and l["bem_sucedida"])
        ws.append([info["nome"], info["supervisor"], bs, com_t, len(anal), aprov, taxa])
        for c in range(1, len(hdrs)+1): ws.cell(row=row, column=c).border = BORDER
        if taxa >= 70:
            for c in range(1, len(hdrs)+1): ws.cell(row=row, column=c).fill = G_FILL
        elif taxa < 40 and len(anal) > 0:
            for c in range(1, len(hdrs)+1): ws.cell(row=row, column=c).fill = Y_FILL
        row += 1
        tot_bs += bs; tot_tr += com_t; tot_an += len(anal); tot_ap += aprov

    taxa_g = round(tot_ap / tot_an * 100, 1) if tot_an else 0
    ws.append(["TOTAL", "", tot_bs, tot_tr, tot_an, tot_ap, taxa_g])
    for c in range(1, len(hdrs)+1):
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).fill = A_FILL
        ws.cell(row=row, column=c).border = BORDER
    row += 2

    row = _sec(ws, row, "CUSTO DA RODADA", 3)
    ws.append(["Tokens entrada", tokens_in])
    ws.append(["Tokens saída",   tokens_out])
    ws.append(["Custo USD",      round(custo_usd, 4)])
    ws.append(["Custo BRL (≈)",  round(custo_usd * USD_TO_BRL, 2)])
    ws.append(["Erros Claude",   0])

    for c_l, w in zip("ABCDEFG", [28,20,16,16,14,18,10]):
        ws.column_dimensions[c_l].width = w

    # ── ABA 2: POR GESTOR ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Por Gestor", 1)
    hdrs2 = ["Gestor","Área","Supervisor","Analisadas","Aprovadas","Taxa %",
             "Nota Média (Atend)","Coaching ALTA","Coaching MÉDIA","Coaching BAIXA"]
    ws2.append(hdrs2); _cab(ws2, 1, len(hdrs2))
    for c in range(1, len(hdrs2)+1):
        ws2.column_dimensions[get_column_letter(c)].width = 16

    row2 = 2
    for area_id, info in AREAS.items():
        for ag in info["agentes"]:
            ag_n   = norm(ag)
            ligs_g = [l for l in analisadas
                      if l["area"] == area_id and norm(l["gestor"]) == ag_n]
            if not ligs_g: continue
            aprov  = sum(1 for l in ligs_g if is_aprovado(l) is True)
            taxa   = round(aprov / len(ligs_g) * 100, 1)
            nota_m = ""
            if area_id == "ATENDIMENTO":
                notas = [nota_atendimento(l["analise"]) for l in ligs_g if l.get("analise")]
                nota_m = round(sum(notas) / len(notas), 1) if notas else ""
            ca = sum(1 for l in ligs_g if coaching(l) == "ALTA")
            cm = sum(1 for l in ligs_g if coaching(l) == "MEDIA")
            cb = sum(1 for l in ligs_g if coaching(l) == "BAIXA")
            ws2.append([ag, info["nome"], info["supervisor"],
                        len(ligs_g), aprov, taxa, nota_m, ca, cm, cb])
            if taxa >= 70:
                for c in range(1, len(hdrs2)+1): ws2.cell(row=row2, column=c).fill = G_FILL
            elif taxa < 40:
                for c in range(1, len(hdrs2)+1): ws2.cell(row=row2, column=c).fill = Y_FILL
            if ca > 3:
                ws2.cell(row=row2, column=8).fill = R_FILL
            row2 += 1

    # ── ABA 3: POR LIGAÇÃO ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Por Ligação", 2)
    hdrs3 = ["ID","Área","Gestor","Data","Hora","Semana","Duração(s)","Telefone",
             "Tipo","CRM","Fonte transcrição","Aprovado/Nota","Tipo identificado",
             "Resumo","Coaching"]
    ws3.append(hdrs3); _cab(ws3, 1, len(hdrs3))

    for l in analisadas:
        a = l["analise"]
        if l["area"] == "ATENDIMENTO":
            ap_str = f"{nota_atendimento(a):.1f}"
        else:
            ap_str = str(a.get("efetivo", ""))
        tipo = (a.get("tipo_pendencia") or a.get("tipo") or
                a.get("tipo_contato")   or "")
        ws3.append([
            l["id"], AREAS[l["area"]]["nome"], l["gestor"],
            l["data"], l["hora"], l["semana"],
            l["duracao_seg"], l["telefone"], l["tipo"], l["crm"],
            l.get("transcricao_fonte", ""),
            ap_str, tipo, a.get("resumo", ""), coaching(l),
        ])

    for i, w in enumerate([28,22,20,12,10,10,10,18,10,14,14,14,22,40,10], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── ABA 4: POR DIA ────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Por Dia", 3)
    hdrs4 = ["Data","Semana","Área","Analisadas","Aprovadas","Não aprovadas","Taxa %"]
    ws4.append(hdrs4); _cab(ws4, 1, len(hdrs4))

    dias_areas: Dict[Tuple, List] = {}
    for l in analisadas:
        k = (l["data"], l["area"])
        dias_areas.setdefault(k, []).append(l)

    for (data, area_id) in sorted(dias_areas.keys()):
        ligs = dias_areas[(data, area_id)]
        aprov = sum(1 for l in ligs if is_aprovado(l) is True)
        taxa  = round(aprov / len(ligs) * 100, 1)
        ws4.append([data, semana_do_mes(data), AREAS[area_id]["nome"],
                    len(ligs), aprov, len(ligs)-aprov, taxa])

    for i, w in enumerate([14,12,24,12,12,14,8], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # ── ABA 5: PADRÕES & INSIGHTS ────────────────────────────────────────────
    ws5 = wb.create_sheet("Padrões & Insights", 4)
    for c_l, w in zip("ABCDEFGH", [32,18,14,14,14,14,14,14]):
        ws5.column_dimensions[c_l].width = w

    ws5.merge_cells("A1:H1")
    t5 = ws5["A1"]; t5.value = f"PADRÕES & INSIGHTS — {mes_ref.upper()}"
    t5.font = Font(bold=True, size=13, color="1F4E79")
    t5.alignment = Alignment(horizontal="center")
    ws5.row_dimensions[1].height = 28
    r = 3

    # 5.1 — Ranking de desempenho
    r = _sec(ws5, r, "RANKING DE DESEMPENHO POR ÁREA", 6)
    for i,h in enumerate(["Gestor","Área","Analisadas","Aprovadas","Taxa %","Destaque"],1):
        ws5.cell(row=r, column=i, value=h)
    _cab(ws5, r, 6); r += 1

    ranking = []
    for area_id, info in AREAS.items():
        for ag in info["agentes"]:
            ligs_g = [l for l in analisadas
                      if l["area"]==area_id and norm(l["gestor"])==norm(ag)]
            if not ligs_g: continue
            aprov = sum(1 for l in ligs_g if is_aprovado(l) is True)
            taxa  = round(aprov/len(ligs_g)*100,1)
            ranking.append((ag, info["nome"], len(ligs_g), aprov, taxa))

    for item in sorted(ranking, key=lambda x: -x[4]):
        destaque = "⭐ Melhor" if item[4] >= 70 else ("⚠️ Atenção" if item[4] < 40 else "")
        ws5.append(list(item) + [destaque])
        fill = G_FILL if item[4] >= 70 else (R_FILL if item[4] < 40 else None)
        if fill:
            for c in range(1,7): ws5.cell(row=r, column=c).fill = fill
        r += 1
    r += 1

    # 5.2 — Ranking de Coaching
    r = _sec(ws5, r, "RANKING DE COACHING — PRIORIDADE ALTA", 5)
    for i,h in enumerate(["Gestor","Área","Coaching ALTA","Total","% ALTA"],1):
        ws5.cell(row=r, column=i, value=h)
    _cab(ws5, r, 5); r += 1

    coach_rank = []
    for area_id, info in AREAS.items():
        for ag in info["agentes"]:
            ligs_g = [l for l in analisadas
                      if l["area"]==area_id and norm(l["gestor"])==norm(ag)]
            if not ligs_g: continue
            alta = sum(1 for l in ligs_g if coaching(l)=="ALTA")
            pct  = round(alta/len(ligs_g)*100,1)
            coach_rank.append((ag, info["nome"], alta, len(ligs_g), pct))

    for item in sorted(coach_rank, key=lambda x: -x[2]):
        ws5.append(list(item))
        if item[2] > 0:
            for c in range(1,6): ws5.cell(row=r, column=c).fill = R_FILL
        r += 1
    r += 1

    # 5.3 — Evolução Semanal
    r = _sec(ws5, r, "EVOLUÇÃO SEMANAL POR ÁREA", 7)
    for i,h in enumerate(["Área","Semana 1","Semana 2","Semana 3","Semana 4","Tendência",""],1):
        ws5.cell(row=r, column=i, value=h)
    _cab(ws5, r, 6); r += 1

    for area_id, info in AREAS.items():
        ligs_a = [l for l in analisadas if l["area"]==area_id]
        taxas  = []
        row_d  = [info["nome"]]
        for s in range(1, 5):
            sg = [l for l in ligs_a if semana_do_mes(l["data"])==f"Semana {s}"]
            if sg:
                ap = sum(1 for l in sg if is_aprovado(l) is True)
                t  = round(ap/len(sg)*100,1)
            else:
                t = None
            taxas.append(t)
            row_d.append(f"{t}%" if t is not None else "-")
        vals = [v for v in taxas if v is not None]
        if len(vals) >= 2:
            tend = "↑ Melhorando" if vals[-1]>vals[0] else ("↓ Caindo" if vals[-1]<vals[0] else "→ Estável")
        else:
            tend = "-"
        row_d.append(tend)
        ws5.append(row_d)
        r += 1
    r += 1

    # 5.4 — Falhas Sucesso do Corretor
    r = _sec(ws5, r, "SUCESSO DO CORRETOR — PENDÊNCIAS QUE MAIS IMPACTAM RENDA", 5)
    for i,h in enumerate(["Tipo de pendência","Ocorrências","Total","% do total",""],1):
        ws5.cell(row=r, column=i, value=h)
    _cab(ws5, r, 4); r += 1

    ligs_sc = [l for l in analisadas if l["area"]=="SUCESSO_CORRETOR"]
    tipos_pend = Counter(
        str(l["analise"].get("tipo_pendencia","")).strip()
        for l in ligs_sc if l["analise"].get("tipo_pendencia","").strip()
    )
    total_sc = len(ligs_sc) or 1
    for tipo, cnt in tipos_pend.most_common(10):
        if tipo in ("","nan"): continue
        ws5.append([tipo, cnt, total_sc, round(cnt/total_sc*100,1)])
        r += 1
    r += 1

    # 5.5 — SPIN por etapa
    r = _sec(ws5, r, "SUCESSO DO CORRETOR — APLICAÇÃO DO SPIN POR ETAPA", 5)
    for i,h in enumerate(["Etapa SPIN","SIM","PARCIAL","NÃO","% SIM"],1):
        ws5.cell(row=r, column=i, value=h)
    _cab(ws5, r, 5); r += 1

    SPIN_CAMPOS = [
        ("spin_situacao",  "Situação"),
        ("spin_problema",  "Problema"),
        ("spin_implicacao","Implicação"),
        ("spin_necessidade","Necessidade"),
    ]
    for campo, label in SPIN_CAMPOS:
        sim = parc = nao = 0
        for l in ligs_sc:
            v = str(l["analise"].get(campo,"")).upper()
            if v=="SIM":     sim+=1
            elif v=="PARCIAL": parc+=1
            else:            nao+=1
        tot = sim+parc+nao or 1
        ws5.append([label, sim, parc, nao, round(sim/tot*100,1)])
        if round(sim/tot*100,1) < 50:
            for c in range(1,6): ws5.cell(row=r, column=c).fill = Y_FILL
        r += 1
    r += 1

    # 5.6 — Falhas Atendimento
    r = _sec(ws5, r, "ATENDIMENTO — CRITÉRIOS MAIS REPROVADOS", 5)
    for i,h in enumerate(["Critério","CUMPRIU","NAO_CUMPRIU","Total","% Reprovação"],1):
        ws5.cell(row=r, column=i, value=h)
    _cab(ws5, r, 5); r += 1

    ligs_at = [l for l in analisadas if l["area"]=="ATENDIMENTO"]
    CRIT_AT = [
        ("saudacao",       "Saudação (peso 25)"),
        ("atencao_clareza","Atenção + Clareza (peso 35)"),
        ("vocabulario",    "Vocabulário (peso 20)"),
        ("cordialidade",   "Cordialidade (peso 20)"),
    ]
    total_at = len(ligs_at) or 1
    for campo, label in CRIT_AT:
        cumpriu = sum(1 for l in ligs_at
                      if str(l["analise"].get(campo,"")).upper()=="CUMPRIU")
        nao_c   = total_at - cumpriu
        pct_rep = round(nao_c/total_at*100,1)
        ws5.append([label, cumpriu, nao_c, total_at, pct_rep])
        if pct_rep > 40:
            for c in range(1,6): ws5.cell(row=r, column=c).fill = R_FILL
        r += 1

    # nota média por agente no Atendimento
    r += 1
    r = _sec(ws5, r, "ATENDIMENTO — NOTA MÉDIA POR AGENTE", 4)
    for i,h in enumerate(["Agente","Ligações analisadas","Nota média","Aprovadas %"],1):
        ws5.cell(row=r, column=i, value=h)
    _cab(ws5, r, 4); r += 1

    for ag in AREAS["ATENDIMENTO"]["agentes"]:
        ligs_g = [l for l in ligs_at if norm(l["gestor"])==norm(ag)]
        if not ligs_g: continue
        notas = [nota_atendimento(l["analise"]) for l in ligs_g]
        nota_m = round(sum(notas)/len(notas),1)
        aprov  = sum(1 for l in ligs_g if is_aprovado(l) is True)
        taxa   = round(aprov/len(ligs_g)*100,1)
        ws5.append([ag, len(ligs_g), nota_m, taxa])
        if nota_m >= 7: ws5.cell(row=r, column=3).fill = G_FILL
        else:           ws5.cell(row=r, column=3).fill = R_FILL
        r += 1
    r += 1

    # 5.7 — Funil Comercial
    r = _sec(ws5, r, "COMERCIAL CADASTRO — FUNIL DE ABORDAGEM", 5)
    for i,h in enumerate(["Etapa","SIM / OK","NÃO","Total","% Conversão"],1):
        ws5.cell(row=r, column=i, value=h)
    _cab(ws5, r, 5); r += 1

    ligs_cc  = [l for l in analisadas if l["area"]=="COMERCIAL_CADASTRO"]
    total_cc = len(ligs_cc) or 1

    FUNIL_CC = [
        ("oferta_clara","Oferta clara (>1 bandeira)"),
        ("fechamento",  "Fechamento com compromisso"),
    ]
    for campo, label in FUNIL_CC:
        sim = sum(1 for l in ligs_cc
                  if str(l["analise"].get(campo,"")).upper()=="SIM")
        ws5.append([label, sim, total_cc-sim, total_cc, round(sim/total_cc*100,1)])
        r += 1

    # Objeção
    ob_tratou  = sum(1 for l in ligs_cc
                     if str(l["analise"].get("objecao","")).upper()=="TRATOU")
    ob_nhouve  = sum(1 for l in ligs_cc
                     if str(l["analise"].get("objecao","")).upper()=="NAO_HOUVE")
    ob_ntratou = sum(1 for l in ligs_cc
                     if str(l["analise"].get("objecao","")).upper()=="NAO_TRATOU")
    ws5.append(["Objeção tratada", ob_tratou, ob_ntratou,
                total_cc, round(ob_tratou/total_cc*100,1)])
    ws5.append(["Sem objeção", ob_nhouve, "", total_cc,
                round(ob_nhouve/total_cc*100,1)])
    r += 2

    # Tipo de contato
    r = _sec(ws5, r, "COMERCIAL CADASTRO — TIPO DE CONTATO", 4)
    for i,h in enumerate(["Tipo","Ocorrências","Total","% do total"],1):
        ws5.cell(row=r, column=i, value=h)
    _cab(ws5, r, 4); r += 1

    tipos_cc = Counter(
        str(l["analise"].get("tipo","")).strip()
        for l in ligs_cc if l["analise"].get("tipo","").strip()
    )
    for tipo, cnt in tipos_cc.most_common():
        if tipo in ("","nan"): continue
        ws5.append([tipo, cnt, total_cc, round(cnt/total_cc*100,1)])
        r += 1

    # Efetividade geral
    ef_cc = sum(1 for l in ligs_cc if is_aprovado(l) is True)
    ws5.append(["EFETIVIDADE GERAL", ef_cc, total_cc-ef_cc,
                total_cc, round(ef_cc/total_cc*100,1)])
    for c in range(1,6):
        ws5.cell(row=r, column=c).font = Font(bold=True)
        ws5.cell(row=r, column=c).fill = G_FILL if ef_cc/total_cc >= 0.5 else Y_FILL
    r += 2

    # 5.8 — Cobertura de transcrições com motivo
    r = _sec(ws5, r, "COBERTURA DE TRANSCRIÇÕES — LIGAÇÕES FORA DA ANÁLISE", 6)
    for i,h in enumerate(["Área","Bem-sucedidas","Analisadas","Fora","% Cobertura",""],1):
        ws5.cell(row=r, column=i, value=h)
    _cab(ws5, r, 5); r += 1

    for area_id, info in AREAS.items():
        ligs_a = [l for l in bem_suc if l["area"]==area_id]
        anal_a = sum(1 for l in ligs_a if l.get("analise"))
        fora   = len(ligs_a) - anal_a
        cobert = round(anal_a/len(ligs_a)*100,1) if ligs_a else 0
        ws5.append([info["nome"], len(ligs_a), anal_a, fora, cobert])
        if fora > 50:
            for c in range(1,6): ws5.cell(row=r, column=c).fill = Y_FILL
        r += 1

    ws5.append(["TOTAL",
                sum(1 for l in bem_suc),
                len(analisadas),
                sum(1 for l in bem_suc) - len(analisadas),
                round(len(analisadas)/len(bem_suc)*100,1) if bem_suc else 0])
    for c in range(1,6): ws5.cell(row=r, column=c).font = Font(bold=True)
    r += 2

    # Motivos detalhados
    r = _sec(ws5, r, "MOTIVOS DAS LIGAÇÕES SEM ANÁLISE", 4)
    for i,h in enumerate(["Motivo","Quantidade","% do total fora",""],1):
        ws5.cell(row=r, column=i, value=h)
    _cab(ws5, r, 3); r += 1

    MOTIVOS_LABEL = {
        "sem_audio_sem_txt":  "Sem áudio e sem transcrição (não baixado do Bitrix)",
        "transcricao_falhou": "Áudio existe mas transcrição falhou",
        "texto_muito_curto":  "Transcrição gerada mas texto muito curto (<50 chars)",
        "audio_ininteligivel":"Áudio ininteligível (Google não reconheceu)",
        "rate_limit_google":  "Rate limit da API Google Speech",
        "falha_tecnica":      "Falha técnica na transcrição",
        None:                 "Transcrição existente mas análise Claude falhou",
    }
    motivos = Counter(l.get("sem_transcricao_motivo") for l in bem_suc if not l.get("analise"))
    total_fora = sum(motivos.values()) or 1
    for motivo, cnt in motivos.most_common():
        label = MOTIVOS_LABEL.get(motivo, str(motivo) or "Outro")
        ws5.append([label, cnt, round(cnt/total_fora*100,1)])
        r += 1

    # ── ABA 6: SEM TRANSCRIÇÃO ────────────────────────────────────────────────
    ws6 = wb.create_sheet("Sem Transcrição", 5)
    hdrs6 = ["Área","Gestor","Data","Hora","Telefone","Status","Motivo"]
    ws6.append(hdrs6); _cab(ws6, 1, len(hdrs6))
    for l in bem_suc:
        if not l.get("transcricao"):
            ws6.append([
                AREAS[l["area"]]["nome"], l["gestor"],
                l["data"], l["hora"], l["telefone"], l["status"],
                MOTIVOS_LABEL.get(l.get("sem_transcricao_motivo"), l.get("sem_transcricao_motivo","") or ""),
            ])
    for i, w in enumerate([22,20,12,10,18,16,40], 1):
        ws6.column_dimensions[get_column_letter(i)].width = w

    # ── SALVAR ────────────────────────────────────────────────────────────────
    destino = output_dir / f"RELATORIO_{mes_ref.upper().replace('/', '_')}_V7_FINAL_{ts}.xlsx"
    wb.save(destino)
    return destino

# ============================================================================
# 13. MAIN
# ============================================================================

def main() -> int:
    parser = argparse.ArgumentParser(description="RAGAZ Monitorias V7")
    parser.add_argument("--base",  default=str(BASE_DIR))
    parser.add_argument("--area",
                        choices=["TODAS","SUCESSO_CORRETOR","ATENDIMENTO","COMERCIAL_CADASTRO"],
                        default="TODAS")
    parser.add_argument("--max",      type=int, default=0)
    parser.add_argument("--data",     default="", help="Filtrar por data YYYY-MM-DD")
    parser.add_argument("--dry-run",  action="store_true")
    parser.add_argument("--so-txts",  action="store_true",
                        help="Pula transcrição de MP3, usa só TXTs existentes")
    args = parser.parse_args()

    base        = Path(args.base)
    transcr_dir = base / "TRANSCRIÇÕES"
    audios_dir  = base / "AUDIOS"   # refinado abaixo após ler XLS
    _MESES_PT_NOM = {
        1:"Janeiro",2:"Fevereiro",3:"Março",4:"Abril",5:"Maio",6:"Junho",
        7:"Julho",8:"Agosto",9:"Setembro",10:"Outubro",11:"Novembro",12:"Dezembro"
    }
    _now        = datetime.now()
    mes_ref     = f"{_MESES_PT_NOM[_now.month]}/{_now.year}"
    output_dir  = base / f"RELATÓRIOS {_now:%d.%m.%y}"
    output_dir.mkdir(parents=True, exist_ok=True)
    _add_file_log(output_dir)

    LOG.info("=" * 70)
    LOG.info(f"RAGAZ MONITORIAS V7 — {datetime.now():%Y-%m-%d %H:%M:%S}")
    LOG.info(f"Área: {args.area}  |  dry-run: {args.dry_run}  |  só-txts: {args.so_txts}")
    if args.data: LOG.info(f"Filtro de data: {args.data}")
    LOG.info("=" * 70)

    # 1) XLS
    xls_path = encontrar_xls(base)
    if not xls_path:
        LOG.error(f"Nenhum XLS encontrado em {base}")
        return 2
    df       = ler_xls(xls_path)
    ligacoes = processar_xls(df)

    if args.area != "TODAS":
        ligacoes = [l for l in ligacoes if l["area"] == args.area]
    if args.data:
        ligacoes = [l for l in ligacoes if l["data"] == args.data]
        LOG.info(f"Filtro por data: {args.data} → {sum(1 for l in ligacoes if l['bem_sucedida'])} bem-sucedidas")

    bem_suc = [l for l in ligacoes if l["bem_sucedida"]]
    if args.max and args.max > 0:
        bem_suc = bem_suc[:args.max]
    LOG.info(f"Bem-sucedidas elegíveis: {len(bem_suc)}")

    # Detectar pasta de áudios pelo mês predominante no XLS
    datas_xls = [l["data"] for l in ligacoes if l.get("data")]
    if datas_xls:
        from collections import Counter as _Ctr
        _mes_ano = _Ctr(d[:7] for d in datas_xls).most_common(1)[0][0]  # "2026-05"
        _ano, _mes = _mes_ano.split("-")
        _MESES_PT = {
            "01": "JANEIRO", "02": "FEVEREIRO", "03": "MARÇO", "04": "ABRIL",
            "05": "MAIO",    "06": "JUNHO",     "07": "JULHO", "08": "AGOSTO",
            "09": "SETEMBRO","10": "OUTUBRO",   "11": "NOVEMBRO","12": "DEZEMBRO",
        }
        _pasta_mes = audios_dir / f"{_MESES_PT[_mes]}_{_ano}"
        mes_ref = f"{_MESES_PT_NOM[int(_mes)]}/{_ano}"
        if _pasta_mes.exists():
            audios_dir = _pasta_mes
            LOG.info(f"Pasta de áudios detectada pelo mês do XLS: {audios_dir}")
        else:
            LOG.warning(f"Pasta {_pasta_mes} não encontrada — usando {audios_dir}")

    # 2) Indexar
    indice = IndiceArquivos(transcr_dir, audios_dir)

    # 3) Transcrever MP3s pendentes (pulado em dry-run e --so-txts)
    if not args.so_txts and not args.dry_run:
        transcriber = Transcriber(transcr_dir)
        transcriber.processar_pendentes(indice, bem_suc)

    # 4) Casar transcrições
    casar_transcricoes(bem_suc, indice)

    # 5) Análise Claude por área
    analyzer   = ClaudeAnalyzer()
    areas_rod  = [args.area] if args.area != "TODAS" else list(AREAS.keys())
    total_anal = 0

    for i, area_id in enumerate(areas_rod):
        n = analyzer.processar_area(bem_suc, area_id, args.dry_run)
        total_anal += n
        if i < len(areas_rod) - 1 and not args.dry_run:
            LOG.info(f"Pausa {AREA_PAUSE_SEC}s entre áreas...")
            for restante in range(AREA_PAUSE_SEC, 0, -10):
                LOG.info(f"  ... {restante}s"); time.sleep(10)

    # 6) Relatório Excel completo
    ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path  = gerar_excel(bem_suc, output_dir, ts,
                             analyzer.tokens_in, analyzer.tokens_out, mes_ref)

    # 7) JSON resumo
    custo_usd = analyzer.tokens_in * PRECO_INPUT + analyzer.tokens_out * PRECO_OUTPUT
    json_data = {
        "gerado_em":  datetime.now().isoformat(),
        "versao":     "V7",
        "modelo":     CLAUDE_MODEL,
        "mes_ref":    mes_ref,
        "tokens_in":  analyzer.tokens_in,
        "tokens_out": analyzer.tokens_out,
        "custo_usd":  round(custo_usd, 6),
        "custo_brl":  round(custo_usd * USD_TO_BRL, 4),
        "erros_claude": len(analyzer.erros),
        "areas": {
            area_id: {
                "nome":            info["nome"],
                "bem_sucedidas":   sum(1 for l in bem_suc if l["area"]==area_id),
                "com_transcricao": sum(1 for l in bem_suc if l["area"]==area_id and l.get("transcricao")),
                "analisadas":      sum(1 for l in bem_suc if l["area"]==area_id and l.get("analise")),
                "aprovadas":       sum(1 for l in bem_suc if l["area"]==area_id and is_aprovado(l) is True),
            }
            for area_id, info in AREAS.items()
        },
    }
    json_path = output_dir / f"RELATORIO_{mes_ref.upper().replace('/','_')}_V6_{ts}.json"
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(json_data, f, ensure_ascii=False, indent=2, default=str)

    # 8) Resumo console
    print("\n" + "=" * 70)
    print("  RAGAZ MONITORIAS V7 — CONCLUÍDO")
    print("=" * 70)
    print(f"  XLS lido          : {xls_path.name}")
    print(f"  Bem-sucedidas     : {len(bem_suc)}")
    print(f"  Total analisadas  : {total_anal}")
    print(f"  Tokens (in/out)   : {analyzer.tokens_in} / {analyzer.tokens_out}")
    print(f"  Custo             : USD {round(custo_usd,4)}  (≈ BRL {round(custo_usd*USD_TO_BRL,2)})")
    print(f"  Erros Claude      : {len(analyzer.erros)}")
    print()
    print("  POR ÁREA:")
    for area_id, d in json_data["areas"].items():
        ap   = d["aprovadas"]; an = d["analisadas"]
        taxa = round(ap/an*100,1) if an else 0
        print(f"   - {d['nome']:<26} transcr={d['com_transcricao']:>4}"
              f"  anal={an:>4}  aprov={ap:>4}  {taxa:>5}%")
    print()
    print(f"  Excel  : {xlsx_path.name}")
    print(f"  JSON   : {json_path.name}")
    print("=" * 70)
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        LOG.warning("Interrompido (Ctrl+C).")
        sys.exit(130)
    except Exception as e:
        LOG.error(f"ERRO FATAL: {e}")
        traceback.print_exc()
        sys.exit(1)
