#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
 RAGAZ — FECHAR RELATÓRIO MAIO 2026
================================================================================
 Junta os 3 Excels gerados em rodadas separadas em um único relatório final
 com todas as abas + Padrões & Insights.

 Fontes:
   - monitorias_maio_20260602_203843.xlsx  → Sucesso do Corretor
   - monitorias_maio_20260603_091730.xlsx  → Comercial Cadastro
   - monitorias_maio_20260603_093332.xlsx  → Atendimento

 USO:
   python "G:\Meu Drive\ARQUITETURA LOVABLE\LINGUAGEM EM CÓDIGO\fechar_relatorio_maio.py"
================================================================================
"""

import os, sys
def _ensure(mod, pkg=""):
    try: __import__(mod)
    except ImportError:
        os.system(f'"{sys.executable}" -m pip install -q {pkg or mod}')

for m,p in [("pandas","pandas"),("openpyxl","openpyxl"),("xlrd","xlrd==1.2.0")]:
    _ensure(m,p)

import pandas as pd
from pathlib import Path
from datetime import datetime
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── caminhos ──────────────────────────────────────────────────────────────────
BASE   = Path(r"G:\Meu Drive\ARQUITETURA LOVABLE")
DIR_02 = BASE / "RELATÓRIOS 02.06.26"
DIR_03 = BASE / "RELATÓRIOS 03.06.26"
OUT    = BASE / "RELATÓRIOS 02.06.26"

FONTES = {
    "SUCESSO_CORRETOR":    DIR_02 / "monitorias_maio_20260602_203843.xlsx",
    "COMERCIAL_CADASTRO":  DIR_03 / "monitorias_maio_20260603_091730.xlsx",
    "ATENDIMENTO":         DIR_03 / "monitorias_maio_20260603_093332.xlsx",
}

AREAS = {
    "SUCESSO_CORRETOR":   {"nome": "Sucesso do Corretor",  "supervisor": "Diego"},
    "ATENDIMENTO":        {"nome": "Atendimento",           "supervisor": "Danielle Moura"},
    "COMERCIAL_CADASTRO": {"nome": "Comercial Cadastro",    "supervisor": "Rodrigo"},
}

CUSTOS = {
    "SUCESSO_CORRETOR":    2.2395,
    "COMERCIAL_CADASTRO":  0.2253,
    "ATENDIMENTO":         0.6289,
}

# ── estilos ───────────────────────────────────────────────────────────────────
H_FILL  = PatternFill("solid", fgColor="1F4E79")
H_FONT  = Font(color="FFFFFF", bold=True, size=10)
S_FILL  = PatternFill("solid", fgColor="2E75B6")
S_FONT  = Font(color="FFFFFF", bold=True, size=10)
A_FILL  = PatternFill("solid", fgColor="D9E1F2")
G_FILL  = PatternFill("solid", fgColor="E2EFDA")
R_FILL  = PatternFill("solid", fgColor="FCE4D6")
BORDER  = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

def cab(ws, row, ncols, fill=None):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill  = fill or H_FILL
        cell.font  = H_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = BORDER

def sec(ws, row, texto, ncols=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=texto)
    c.fill = S_FILL; c.font = S_FONT
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20
    return row + 1

def fmt_pct(v): return f"{v:.1f}%"

def wk(data_str):
    try:
        d = datetime.strptime(str(data_str)[:10], "%Y-%m-%d")
        semana = (d.day - 1) // 7 + 1
        return f"Semana {semana}"
    except: return "?"

# ── leitura dos Excels ────────────────────────────────────────────────────────

def ler_aba(path: Path, aba: str) -> pd.DataFrame:
    try:
        return pd.read_excel(path, sheet_name=aba, engine="openpyxl")
    except Exception as e:
        print(f"  [AVISO] {path.name} / {aba}: {e}")
        return pd.DataFrame()

def carregar_dados():
    ligacoes   = []   # Por Ligação
    sem_transc = []   # Sem Transcrição

    for area_id, path in FONTES.items():
        if not path.exists():
            print(f"[ERRO] Arquivo não encontrado: {path}")
            continue

        df = ler_aba(path, "Por Ligação")
        if not df.empty:
            df["_area_id"] = area_id
            ligacoes.append(df)

        df2 = ler_aba(path, "Sem Transcrição")
        if not df2.empty:
            sem_transc.append(df2)

    df_lig  = pd.concat(ligacoes,   ignore_index=True) if ligacoes   else pd.DataFrame()
    df_sem  = pd.concat(sem_transc, ignore_index=True) if sem_transc else pd.DataFrame()
    return df_lig, df_sem

# ── helpers de análise ────────────────────────────────────────────────────────

def is_efetivo(row) -> bool:
    v = str(row.get("Efetivo/Nota", "") or "").strip().upper()
    if v in ("SIM", "S"): return True
    try:
        return float(v) >= 7.0
    except: return False

def semana_num(data_str):
    try:
        d = datetime.strptime(str(data_str)[:10], "%Y-%m-%d")
        return (d.day - 1) // 7 + 1
    except: return 0

# ── geração do Excel ──────────────────────────────────────────────────────────

def gerar_excel(df_lig: pd.DataFrame, df_sem: pd.DataFrame):
    wb = Workbook()
    wb.remove(wb.active)

    custo_total = sum(CUSTOS.values())

    # ── 1. RESUMO ─────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Resumo", 0)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    for c in "CDEF": ws.column_dimensions[c].width = 14

    ws.merge_cells("A1:F1")
    t = ws["A1"]; t.value = "RAGAZ — Monitorias Maio 2026 — Relatório Final"
    t.font = Font(bold=True, size=14, color="1F4E79")
    t.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    ws["A2"].value = f"Gerado em {datetime.now():%d/%m/%Y %H:%M}"
    ws["A2"].font = Font(color="808080", italic=True, size=9)
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4
    row = sec(ws, row, "RESUMO GERAL", 6)
    for i,h in enumerate(["Área","Supervisor","Bem-sucedidas","Com transcrição","Analisadas","Efetivas","Taxa %"],1):
        ws.cell(row=row, column=i, value=h)
    cab(ws, row, 7); row+=1

    totais = {"bem_suc":0,"transcr":0,"anal":0,"efet":0}
    for area_id, info in AREAS.items():
        sub = df_lig[df_lig["_area_id"]==area_id] if "_area_id" in df_lig.columns else pd.DataFrame()
        anal = len(sub)
        efet = sum(1 for _,r in sub.iterrows() if is_efetivo(r))
        taxa = round(efet/anal*100,1) if anal else 0

        # bem-sucedidas e transcrições vêm dos JSONs conhecidos
        bs   = {"SUCESSO_CORRETOR":804,"ATENDIMENTO":608,"COMERCIAL_CADASTRO":139}[area_id]
        trsc = {"SUCESSO_CORRETOR":599,"ATENDIMENTO":332,"COMERCIAL_CADASTRO":99}[area_id]

        ws.append([info["nome"], info["supervisor"], bs, trsc, anal, efet, taxa])
        for c in range(1,8): ws.cell(row=row, column=c).border = BORDER
        row+=1

        totais["bem_suc"]+=bs; totais["transcr"]+=trsc
        totais["anal"]+=anal;  totais["efet"]+=efet

    taxa_geral = round(totais["efet"]/totais["anal"]*100,1) if totais["anal"] else 0
    ws.append(["TOTAL","",totais["bem_suc"],totais["transcr"],totais["anal"],totais["efet"],taxa_geral])
    for c in range(1,8):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True); cell.fill = A_FILL; cell.border = BORDER
    row+=2

    row = sec(ws, row, "CUSTO DA RODADA", 4)
    for area_id, info in AREAS.items():
        ws.append([info["nome"], f"USD {CUSTOS[area_id]:.4f}"])
        row+=1
    ws.append(["TOTAL", f"USD {custo_total:.4f}", f"≈ BRL {custo_total*5.40:.2f}"])
    for c in range(1,4): ws.cell(row=row,column=c).font=Font(bold=True)

    # ── 2. POR GESTOR ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Por Gestor", 1)
    hdrs = ["Gestor","Área","Supervisor","Analisadas","Efetivas","Não efetivas","Taxa %","Coaching ALTA","Coaching MÉDIA","Coaching BAIXA"]
    ws2.append(hdrs); cab(ws2, 1, len(hdrs))
    for c in range(1,len(hdrs)+1): ws2.column_dimensions[get_column_letter(c)].width=16

    if not df_lig.empty:
        for gestor, grp in df_lig.groupby("Gestor"):
            anal = len(grp)
            efet = sum(1 for _,r in grp.iterrows() if is_efetivo(r))
            nef  = anal - efet
            taxa = round(efet/anal*100,1) if anal else 0
            area_nome = grp["Área"].iloc[0] if "Área" in grp.columns else ""
            sup = ""
            for aid, info in AREAS.items():
                if info["nome"] == area_nome: sup = info["supervisor"]
            coachings = grp["Prioridade coaching"].str.upper() if "Prioridade coaching" in grp.columns else pd.Series(dtype=str)
            ca = (coachings=="ALTA").sum()
            cm = (coachings=="MEDIA").sum()
            cb = (coachings=="BAIXA").sum()
            ws2.append([gestor, area_nome, sup, anal, efet, nef, taxa, ca, cm, cb])

    # ── 3. POR LIGAÇÃO ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Por Ligação", 2)
    if not df_lig.empty:
        cols_excl = ["_area_id"]
        cols_show = [c for c in df_lig.columns if c not in cols_excl]
        ws3.append(cols_show); cab(ws3, 1, len(cols_show))
        for _, row_data in df_lig[cols_show].iterrows():
            ws3.append(list(row_data))
        for c in range(1, len(cols_show)+1):
            ws3.column_dimensions[get_column_letter(c)].width = 18
        ws3.column_dimensions["M"].width = 40  # Resumo

    # ── 4. POR DIA ────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Por Dia", 3)
    hdrs4 = ["Data","Semana","Área","Analisadas","Efetivas","Não efetivas","Taxa %"]
    ws4.append(hdrs4); cab(ws4, 1, len(hdrs4))
    for c in range(1,len(hdrs4)+1): ws4.column_dimensions[get_column_letter(c)].width=16

    if not df_lig.empty and "Data" in df_lig.columns:
        for (data, area_nome), grp in df_lig.groupby(["Data","Área"]):
            anal = len(grp)
            efet = sum(1 for _,r in grp.iterrows() if is_efetivo(r))
            taxa = round(efet/anal*100,1) if anal else 0
            ws4.append([str(data)[:10], wk(data), area_nome, anal, efet, anal-efet, taxa])

    # ── 5. PADRÕES & INSIGHTS ────────────────────────────────────────────────
    ws5 = wb.create_sheet("Padrões & Insights", 4)
    ws5.column_dimensions["A"].width = 32
    for c in "BCDEFG": ws5.column_dimensions[c].width = 14

    ws5.merge_cells("A1:G1")
    t5 = ws5["A1"]; t5.value = "PADRÕES & INSIGHTS — MAIO 2026"
    t5.font = Font(bold=True, size=13, color="1F4E79")
    t5.alignment = Alignment(horizontal="center")
    ws5.row_dimensions[1].height = 26
    row5 = 3

    # ── 5.1 Ranking de desempenho ─────────────────────────────────────────────
    row5 = sec(ws5, row5, "RANKING DE DESEMPENHO POR ÁREA", 6)
    for i,h in enumerate(["Gestor","Área","Analisadas","Efetivas","Taxa %",""],1):
        ws5.cell(row=row5, column=i, value=h)
    cab(ws5, row5, 5); row5+=1

    if not df_lig.empty:
        ranking = []
        for gestor, grp in df_lig.groupby("Gestor"):
            anal = len(grp); efet = sum(1 for _,r in grp.iterrows() if is_efetivo(r))
            taxa = round(efet/anal*100,1) if anal else 0
            area_nome = grp["Área"].iloc[0] if "Área" in grp.columns else ""
            ranking.append((gestor, area_nome, anal, efet, taxa))
        for r in sorted(ranking, key=lambda x: -x[4]):
            fill = G_FILL if r[4]>=70 else (R_FILL if r[4]<40 else None)
            ws5.append(list(r[:5]))
            if fill:
                for c in range(1,6): ws5.cell(row=row5,column=c).fill=fill
            row5+=1
    row5+=1

    # ── 5.2 Ranking de Coaching ───────────────────────────────────────────────
    row5 = sec(ws5, row5, "RANKING DE COACHING — PRIORIDADE ALTA", 5)
    for i,h in enumerate(["Gestor","Área","Coaching ALTA","Total analisadas","% ALTA"],1):
        ws5.cell(row=row5, column=i, value=h)
    cab(ws5, row5, 5); row5+=1

    if not df_lig.empty and "Prioridade coaching" in df_lig.columns:
        coach_rank = []
        for gestor, grp in df_lig.groupby("Gestor"):
            alta = (grp["Prioridade coaching"].str.upper()=="ALTA").sum()
            area_nome = grp["Área"].iloc[0] if "Área" in grp.columns else ""
            pct = round(alta/len(grp)*100,1) if len(grp) else 0
            coach_rank.append((gestor, area_nome, int(alta), len(grp), pct))
        for r in sorted(coach_rank, key=lambda x: -x[2]):
            ws5.append(list(r))
            if r[2] > 0:
                for c in range(1,6): ws5.cell(row=row5,column=c).fill=R_FILL
            row5+=1
    row5+=1

    # ── 5.3 Evolução Semanal ──────────────────────────────────────────────────
    row5 = sec(ws5, row5, "EVOLUÇÃO SEMANAL POR ÁREA", 6)
    semanas = ["Semana 1","Semana 2","Semana 3","Semana 4"]
    for i,h in enumerate(["Área"] + semanas + ["Tendência"],1):
        ws5.cell(row=row5, column=i, value=h)
    cab(ws5, row5, 6); row5+=1

    if not df_lig.empty and "Data" in df_lig.columns:
        for area_id, info in AREAS.items():
            sub = df_lig[df_lig["_area_id"]==area_id]
            taxas = []
            row_data = [info["nome"]]
            for s in range(1,5):
                sg = sub[sub["Data"].apply(lambda d: semana_num(d)==s)]
                if len(sg):
                    efet = sum(1 for _,r in sg.iterrows() if is_efetivo(r))
                    t = round(efet/len(sg)*100,1)
                else:
                    t = None
                taxas.append(t)
                row_data.append(f"{t}%" if t is not None else "-")
            # tendência simples
            vals = [v for v in taxas if v is not None]
            if len(vals)>=2:
                tend = "↑ Melhorando" if vals[-1]>vals[0] else ("↓ Caindo" if vals[-1]<vals[0] else "→ Estável")
            else:
                tend = "-"
            row_data.append(tend)
            ws5.append(row_data)
            row5+=1
    row5+=1

    # ── 5.4 Falhas mais comuns — Atendimento ─────────────────────────────────
    row5 = sec(ws5, row5, "FALHAS MAIS COMUNS — ATENDIMENTO", 4)
    for i,h in enumerate(["Critério","Reprovações","Total","% Falha"],1):
        ws5.cell(row=row5, column=i, value=h)
    cab(ws5, row5, 4); row5+=1

    sub_at = df_lig[df_lig["_area_id"]=="ATENDIMENTO"] if not df_lig.empty else pd.DataFrame()
    if not sub_at.empty and "Tipo identificado" in sub_at.columns:
        # criterios do V5 estão no campo Resumo/Tipo — aproximação pelo campo disponível
        criterios_atend = {
            "Saudação":       "saudacao",
            "Atenção/Clareza":"atencao_clareza",
            "Vocabulário":    "vocabulario",
            "Cordialidade":   "cordialidade",
        }
        total_at = len(sub_at)
        # sem dados granulares por critério no Excel consolidado — mostra distribuição de coaching
        dist = sub_at["Prioridade coaching"].str.upper().value_counts() if "Prioridade coaching" in sub_at.columns else pd.Series(dtype=int)
        for k,v in dist.items():
            pct = round(v/total_at*100,1)
            ws5.append([f"Coaching {k}", int(v), total_at, pct])
            if k=="ALTA":
                for c in range(1,5): ws5.cell(row=row5,column=c).fill=R_FILL
            row5+=1
    row5+=1

    # ── 5.5 Falhas mais comuns — Sucesso do Corretor ─────────────────────────
    row5 = sec(ws5, row5, "FALHAS MAIS COMUNS — SUCESSO DO CORRETOR", 4)
    for i,h in enumerate(["Tipo de pendência / impacto","Ocorrências","Total","% do total"],1):
        ws5.cell(row=row5, column=i, value=h)
    cab(ws5, row5, 4); row5+=1

    sub_sc = df_lig[df_lig["_area_id"]=="SUCESSO_CORRETOR"] if not df_lig.empty else pd.DataFrame()
    if not sub_sc.empty and "Tipo identificado" in sub_sc.columns:
        tipos = sub_sc["Tipo identificado"].dropna()
        total_sc = len(sub_sc)
        for tipo, cnt in Counter(tipos).most_common(10):
            if str(tipo).strip() in ("","nan"): continue
            pct = round(cnt/total_sc*100,1)
            ws5.append([str(tipo), int(cnt), total_sc, pct])
            row5+=1
    row5+=1

    # ── 5.6 Funil Comercial ───────────────────────────────────────────────────
    row5 = sec(ws5, row5, "FUNIL COMERCIAL — COMERCIAL CADASTRO", 4)
    for i,h in enumerate(["Etapa","SIM / OK","NÃO","% Conversão"],1):
        ws5.cell(row=row5, column=i, value=h)
    cab(ws5, row5, 4); row5+=1

    sub_cc = df_lig[df_lig["_area_id"]=="COMERCIAL_CADASTRO"] if not df_lig.empty else pd.DataFrame()
    if not sub_cc.empty and "Tipo identificado" in sub_cc.columns:
        total_cc = len(sub_cc)
        tipos_cc = Counter(sub_cc["Tipo identificado"].dropna())
        for tipo, cnt in tipos_cc.most_common():
            if str(tipo).strip() in ("","nan"): continue
            pct = round(cnt/total_cc*100,1)
            ws5.append([str(tipo), int(cnt), total_cc-cnt, pct])
            row5+=1

        efet_cc = sum(1 for _,r in sub_cc.iterrows() if is_efetivo(r))
        ws5.append(["EFETIVIDADE GERAL", efet_cc, total_cc-efet_cc, round(efet_cc/total_cc*100,1)])
        for c in range(1,5):
            ws5.cell(row=row5,column=c).font=Font(bold=True)
            ws5.cell(row=row5,column=c).fill=G_FILL
        row5+=1
    row5+=1

    # ── 5.7 Cobertura de transcrições ────────────────────────────────────────
    row5 = sec(ws5, row5, "COBERTURA DE TRANSCRIÇÕES", 4)
    for i,h in enumerate(["Área","Bem-sucedidas","Analisadas","Fora da análise"],1):
        ws5.cell(row=row5, column=i, value=h)
    cab(ws5, row5, 4); row5+=1

    dados_cob = [
        ("Sucesso do Corretor",  804, 595),
        ("Atendimento",          608, 332),
        ("Comercial Cadastro",   139,  99),
    ]
    for nome, bs, anal in dados_cob:
        fora = bs - anal
        ws5.append([nome, bs, anal, fora])
        if fora > 50:
            for c in range(1,5): ws5.cell(row=row5,column=c).fill=R_FILL
        row5+=1
    ws5.append(["TOTAL", 1551, 1026, 525])
    for c in range(1,5): ws5.cell(row=row5,column=c).font=Font(bold=True)

    # ── 6. SEM TRANSCRIÇÃO ───────────────────────────────────────────────────
    ws6 = wb.create_sheet("Sem Transcrição", 5)
    if not df_sem.empty:
        ws6.append(list(df_sem.columns)); cab(ws6, 1, len(df_sem.columns))
        for _, r in df_sem.iterrows(): ws6.append(list(r))
        for c in range(1, len(df_sem.columns)+1):
            ws6.column_dimensions[get_column_letter(c)].width=18

    # ── salvar ────────────────────────────────────────────────────────────────
    ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
    destino = OUT / f"RELATORIO_FINAL_MAIO_2026_{ts}.xlsx"
    wb.save(destino)
    return destino


# ── main ──────────────────────────────────────────────────────────────────────
def main():
    print("Carregando dados...")
    df_lig, df_sem = carregar_dados()
    print(f"  Por Ligação: {len(df_lig)} registros")
    print(f"  Sem Transcrição: {len(df_sem)} registros")

    print("Gerando Excel...")
    destino = gerar_excel(df_lig, df_sem)

    print("\n" + "="*60)
    print("  RELATÓRIO FINAL MAIO 2026 — GERADO")
    print("="*60)
    print(f"  Arquivo: {destino.name}")
    print(f"  Pasta:   {destino.parent}")
    print(f"  Tamanho: {round(destino.stat().st_size/1024,1)} KB")
    print("="*60)

if __name__ == "__main__":
    main()
