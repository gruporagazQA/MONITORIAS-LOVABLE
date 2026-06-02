#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
 RAGAZ — MONITORIAS 2026 — V3
================================================================================
 Novidades em relação à V2:
   - Transcreve MP3s que ainda não têm TXT correspondente
   - Salva novos TXTs em TRANSCRIÇÕES\ para reuso futuro
   - Matching áudio/transcrição por data + últimos 8 dígitos do telefone
   - Mantém: delay 1.5s entre chamadas, pausa 65s entre áreas, bug Ariely OK

 USO (PowerShell):
   python "G:\Meu Drive\ARQUITETURA LOVABLE\LINGUAGEM EM CÓDIGO\V3_monitorias_ragaz.py"

 OPÇÕES:
   --area     TODAS | SUCESSO_CORRETOR | ATENDIMENTO | COMERCIAL_CADASTRO
   --max      N      (limita N ligações — útil para teste)
   --dry-run         (sem Claude: só transcreve e casa, gera Excel parcial)
   --so-txts         (pula transcrição de MP3 — usa só TXTs existentes)

 PRÉ-REQUISITO:
   $env:ANTHROPIC_API_KEY = "sk-ant-..."
================================================================================
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
import tempfile
import time
import traceback
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# ============================================================================
# 0. DEPENDÊNCIAS
# ============================================================================

def _ensure(module: str, package: str = "") -> None:
    try:
        __import__(module)
    except ImportError:
        pkg = package or module
        print(f"[setup] instalando {pkg} ...", flush=True)
        os.system(f'"{sys.executable}" -m pip install -q {pkg}')

for mod, pkg in [
    ("pandas",           "pandas"),
    ("openpyxl",         "openpyxl"),
    ("xlrd",             "xlrd==1.2.0"),
    ("lxml",             "lxml"),
    ("anthropic",        "anthropic"),
    ("tqdm",             "tqdm"),
    ("librosa",          "librosa"),
    ("soundfile",        "soundfile"),
    ("scipy",            "scipy"),
    ("speech_recognition","SpeechRecognition"),
]:
    _ensure(mod, pkg)

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from tqdm import tqdm

# ============================================================================
# 1. CONFIGURAÇÃO
# ============================================================================

BASE_DIR = Path(r"G:\Meu Drive\ARQUITETURA LOVABLE")

AREAS: Dict[str, Dict[str, Any]] = {
    "SUCESSO_CORRETOR": {
        "nome":       "Sucesso do Corretor",
        "supervisor": "Diego",
        "agentes":    ["Aline De Souza", "Ariel Maia", "Tatiana Costa",
                       "Taynnã Alves", "Ligiane Maia", "Eduardo Novais"],
    },
    "ATENDIMENTO": {
        "nome":       "Atendimento",
        "supervisor": "Danielle Moura",
        "agentes":    ["Deuzeni", "Lhaine", "Talita", "Jéssica"],
    },
    "COMERCIAL_CADASTRO": {
        "nome":       "Comercial Cadastro",
        "supervisor": "Rodrigo",
        "agentes":    ["Natália", "Ariely", "Ana Lúcia"],
    },
}

# Claude
CLAUDE_MODEL        = "claude-haiku-4-5-20251001"
MAX_TOKENS_ANALISE  = 700
MAX_PALAVRAS_TRANSC = 1500
MAX_RETRIES_CLAUDE  = 3
CLAUDE_RETRY_DELAY  = 4

# Rate limit
CALL_DELAY_SEC = 1.5
AREA_PAUSE_SEC = 65

# Preços
PRECO_INPUT  = 0.80  / 1_000_000
PRECO_OUTPUT = 4.00  / 1_000_000
USD_TO_BRL   = 5.40

# XLS
COLUNAS_DATA  = ["Data da chamada", "Data", "Data/Hora"]
COLUNAS_DUR   = ["Duração da chamada", "Duraçãoda chamada", "Duração", "Duracao"]
COLUNAS_TIPO  = ["Tipo Chamada", "Tipo de chamada", "Tipo"]
COLUNAS_CRM   = ["CRM", "Contato CRM"]
COLUNAS_OBS   = ["Observação", "Observacao"]
TIPO_EF       = ("efetuad", "saída", "saida", "outbound")
TIPO_REC      = ("recebid", "entrada", "inbound")
STATUS_OK     = "Bem sucedida"
DUR_MIN_SEG   = 30

# Áudio
AUDIO_SR      = 16000
SPEECH_LANG   = "pt-BR"

# ============================================================================
# 2. LOGGING
# ============================================================================

LOG = logging.getLogger("ragaz_v3")
LOG.setLevel(logging.INFO)
_fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%Y-%m-%d %H:%M:%S")
_sh  = logging.StreamHandler(sys.stdout)
_sh.setFormatter(_fmt)
LOG.addHandler(_sh)

def _add_file_log(output_dir: Path) -> None:
    (output_dir / "logs").mkdir(parents=True, exist_ok=True)
    fh = logging.FileHandler(
        output_dir / "logs" / f"rodada_{datetime.now():%Y%m%d_%H%M%S}.log",
        encoding="utf-8"
    )
    fh.setFormatter(_fmt)
    LOG.addHandler(fh)

# ============================================================================
# 3. UTILITÁRIOS
# ============================================================================

def norm(s: str) -> str:
    if not s:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(s))
    return "".join(c for c in nfkd if not unicodedata.combining(c)).lower().strip()


def normalizar_tel(t: Any) -> str:
    if t is None or str(t).strip() in ("", "-", "nan"):
        return ""
    digits = "".join(c for c in str(t) if c.isdigit())
    if not digits:
        return ""
    if not digits.startswith("55") and len(digits) >= 10:
        digits = "55" + digits[-10:]
    return "+" + digits


def u8(tel: str) -> str:
    """Últimos 8 dígitos do telefone (sem +)."""
    digits = tel.replace("+", "")
    return digits[-8:] if len(digits) >= 8 else digits


def dur_seg(d: Any) -> int:
    if d is None or str(d).strip() in ("", "-", "nan"):
        return 0
    s = str(d).strip().lower()
    seg = 0
    m = re.search(r"(\d+)\s*m", s)
    if m:
        seg += int(m.group(1)) * 60
    m = re.search(r"(\d+)\s*s(?![a-z])", s)
    if m:
        seg += int(m.group(1))
    if seg == 0:
        raw = re.sub(r"\D", "", s)
        seg = int(raw) if raw else 0
    return seg


def parse_dt(s: Any) -> Optional[datetime]:
    if s is None:
        return None
    txt = str(s).strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S",
                "%d/%m/%Y %H:%M",    "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(txt, fmt)
        except ValueError:
            continue
    return None


def truncar(texto: str, n: int = MAX_PALAVRAS_TRANSC) -> str:
    palavras = texto.split()
    return texto if len(palavras) <= n else " ".join(palavras[:n]) + " [...truncado...]"


def classificar_agente(nome: str) -> Optional[str]:
    """
    Corrigido: match exato primeiro, depois parcial (evita bug Ariely/Ariel).
    """
    if not nome:
        return None
    alvo = norm(nome)
    for area_id, info in AREAS.items():
        for ag in info["agentes"]:
            if norm(ag) == alvo:
                return area_id
    for area_id, info in AREAS.items():
        for ag in info["agentes"]:
            ag_n = norm(ag)
            if len(ag_n) >= 4 and len(alvo) >= 4:
                if ag_n in alvo or alvo in ag_n:
                    return area_id
    return None


def col(df: pd.DataFrame, candidatas: List[str]) -> Optional[str]:
    cols_n = {norm(c): c for c in df.columns}
    for cand in candidatas:
        n = norm(cand)
        if n in cols_n:
            return cols_n[n]
    for cand in candidatas:
        n = norm(cand)
        for k, real in cols_n.items():
            if n in k:
                return real
    return None

# ============================================================================
# 4. LEITURA DO XLS
# ============================================================================

def encontrar_xls(base: Path) -> Optional[Path]:
    candidatos = [*base.glob("calls_detail_*.xlsx"), *base.glob("calls_detail_*.xls"),
                  *base.glob("*.xlsx"), *base.glob("*.xls")]
    candidatos = [c for c in candidatos if c.is_file()]
    return max(candidatos, key=lambda p: p.stat().st_mtime) if candidatos else None


def ler_xls(arquivo: Path) -> pd.DataFrame:
    LOG.info(f"XLS: {arquivo.name}")
    try:
        return pd.read_excel(arquivo, engine="openpyxl" if arquivo.suffix == ".xlsx" else "xlrd")
    except Exception as e:
        LOG.warning(f"read_excel falhou ({e}), tentando read_html...")
        tabelas = pd.read_html(str(arquivo))
        if not tabelas:
            raise RuntimeError("Nenhuma tabela encontrada no XLS")
        return tabelas[0]


def processar_xls(df: pd.DataFrame) -> List[Dict[str, Any]]:
    c_col  = col(df, ["Colaborador"])
    c_tel  = col(df, ["Telefone"])
    c_data = col(df, COLUNAS_DATA)
    c_dur  = col(df, COLUNAS_DUR)
    c_tipo = col(df, COLUNAS_TIPO)
    c_stat = col(df, ["Status"])
    c_crm  = col(df, COLUNAS_CRM)
    c_obs  = col(df, COLUNAS_OBS)

    faltando = [nm for nm, c in [("Colaborador", c_col), ("Telefone", c_tel),
                                  ("Data", c_data), ("Status", c_stat)] if c is None]
    if faltando:
        raise RuntimeError(f"Colunas obrigatórias ausentes: {faltando}")

    ligacoes = []
    for idx, row in df.iterrows():
        gestor  = str(row.get(c_col, "")).strip()
        tel     = normalizar_tel(row.get(c_tel, ""))
        dt      = parse_dt(row.get(c_data, ""))
        dur_s   = dur_seg(row.get(c_dur, "")) if c_dur else 0
        tipo_r  = norm(row.get(c_tipo, "")) if c_tipo else ""
        status  = str(row.get(c_stat, "")).strip()
        crm     = str(row.get(c_crm, "")).strip() if c_crm else ""
        obs     = str(row.get(c_obs, "")).strip() if c_obs else ""

        is_ef   = any(t in tipo_r for t in TIPO_EF)
        is_rec  = any(t in tipo_r for t in TIPO_REC)
        is_ok   = (status == STATUS_OK and dur_s > DUR_MIN_SEG)
        area_id = classificar_agente(gestor)
        if area_id is None:
            continue

        ligacoes.append({
            "area":        area_id,
            "gestor":      gestor,
            "telefone":    tel,
            "datetime":    dt,
            "data":        dt.strftime("%Y-%m-%d") if dt else "",
            "hora":        dt.strftime("%H:%M:%S") if dt else "",
            "duracao_seg": dur_s,
            "tipo":        "efetuada" if is_ef else ("recebida" if is_rec else "outro"),
            "status":      status,
            "crm":         crm,
            "obs":         obs,
            "bem_sucedida":is_ok,
            "transcricao": None,
            "transcricao_fonte": None,
            "analise":     None,
        })

    LOG.info(f"XLS: {len(df)} linhas → {len(ligacoes)} em escopo")
    return ligacoes

# ============================================================================
# 5. ÍNDICE DE TRANSCRIÇÕES E ÁUDIOS
# ============================================================================

class IndiceArquivos:
    """
    Indexa TXTs e MP3s por chave (data + últimos 8 dígitos do telefone).
    Áudio: 'YYYY-MM-DD HH-MM-SS +?PHONE.mp3'
    TXT:   'YYYY-MM-DD_HH-MM-SS_+?PHONE.txt'
    """

    PAT_TXT = re.compile(r"^(\d{4}-\d{2}-\d{2})_[\d-]+_\+?(\d+)\.txt$")
    PAT_MP3 = re.compile(r"^(\d{4}-\d{2}-\d{2}) [\d-]+ \+?(\d+)\.mp3$")

    def __init__(self, transcr_dir: Path, audios_dir: Path) -> None:
        self.transcr_dir = transcr_dir
        self.audios_dir  = audios_dir
        self.txts:  Dict[str, Path] = {}   # chave → Path .txt
        self.mp3s:  Dict[str, Path] = {}   # chave → Path .mp3
        self._indexar()

    def _chave(self, data: str, tel_digits: str) -> str:
        return f"{data}|{u8(tel_digits)}"

    def _indexar(self) -> None:
        if self.transcr_dir.exists():
            for arq in self.transcr_dir.glob("*.txt"):
                m = self.PAT_TXT.match(arq.name)
                if m:
                    self.txts[self._chave(m.group(1), m.group(2))] = arq
        LOG.info(f"TXTs indexados: {len(self.txts)}")

        if self.audios_dir.exists():
            for arq in self.audios_dir.rglob("*.mp3"):
                m = self.PAT_MP3.match(arq.name)
                if m:
                    self.mp3s[self._chave(m.group(1), m.group(2))] = arq
        LOG.info(f"MP3s indexados: {len(self.mp3s)}")

    def buscar_txt(self, data: str, tel: str) -> Optional[Path]:
        return self.txts.get(self._chave(data, tel.replace("+", "")))

    def buscar_mp3(self, data: str, tel: str) -> Optional[Path]:
        return self.mp3s.get(self._chave(data, tel.replace("+", "")))

# ============================================================================
# 6. TRANSCRIÇÃO DE ÁUDIO
# ============================================================================

class Transcriber:
    def __init__(self, transcr_dir: Path) -> None:
        self.dir = transcr_dir
        self.dir.mkdir(parents=True, exist_ok=True)
        import speech_recognition as sr
        self._sr = sr
        self.rec = sr.Recognizer()

    def transcrever_mp3(self, mp3: Path) -> Tuple[Optional[str], Optional[str]]:
        """Retorna (texto, erro)."""
        try:
            import librosa
            import scipy.io.wavfile as wavfile
            import numpy as np

            audio_data, sr_audio = librosa.load(str(mp3), sr=AUDIO_SR, mono=True)
            with tempfile.NamedTemporaryFile(suffix=".wav", delete=False) as tmp:
                tmp_path = tmp.name
            wavfile.write(tmp_path, sr_audio, (audio_data * 32767).astype(np.int16))

            try:
                with self._sr.AudioFile(tmp_path) as source:
                    audio = self.rec.record(source)
                texto = self.rec.recognize_google(audio, language=SPEECH_LANG)
                return texto, None
            except self._sr.UnknownValueError:
                return None, "Áudio ininteligível"
            except self._sr.RequestError as e:
                msg = str(e)
                if "403" in msg or "Forbidden" in msg:
                    return None, "Rate limit Google API"
                return None, f"Erro Google API: {msg[:80]}"
            finally:
                try:
                    Path(tmp_path).unlink(missing_ok=True)
                except Exception:
                    pass
        except Exception as e:
            return None, f"Falha: {str(e)[:100]}"

    def nome_txt(self, mp3: Path) -> str:
        """Converte nome do MP3 para nome do TXT (espaço → underscore)."""
        return mp3.stem.replace(" ", "_") + ".txt"

    def salvar(self, mp3: Path, texto: str) -> Path:
        destino = self.dir / self.nome_txt(mp3)
        destino.write_text(texto, encoding="utf-8")
        return destino

    def processar_pendentes(self, indice: IndiceArquivos, ligacoes: List[Dict]) -> int:
        """Transcreve MP3s de ligações que não têm TXT. Salva e atualiza índice."""
        pendentes = [
            l for l in ligacoes
            if l["bem_sucedida"]
            and indice.buscar_txt(l["data"], l["telefone"]) is None
            and indice.buscar_mp3(l["data"], l["telefone"]) is not None
        ]
        if not pendentes:
            LOG.info("Nenhum MP3 pendente de transcrição.")
            return 0

        LOG.info(f"Transcrevendo {len(pendentes)} MP3s sem TXT...")
        ok = 0
        for lig in tqdm(pendentes, desc="Transcrição MP3", unit="áudio"):
            mp3 = indice.buscar_mp3(lig["data"], lig["telefone"])
            texto, erro = self.transcrever_mp3(mp3)
            if erro or not texto:
                LOG.debug(f"  {mp3.name}: {erro}")
                continue
            destino = self.salvar(mp3, texto)
            # Atualiza índice em memória
            chave = indice._chave(lig["data"], lig["telefone"].replace("+", ""))
            indice.txts[chave] = destino
            ok += 1

        LOG.info(f"Transcrições novas geradas: {ok}/{len(pendentes)}")
        return ok

# ============================================================================
# 7. MATCH LIGAÇÃO ↔ TRANSCRIÇÃO
# ============================================================================

def casar_transcricoes(ligacoes: List[Dict], indice: IndiceArquivos) -> int:
    encontradas = 0
    for lig in ligacoes:
        if not lig["bem_sucedida"]:
            continue
        arq = indice.buscar_txt(lig["data"], lig["telefone"])
        if arq:
            txt = arq.read_text(encoding="utf-8", errors="replace").strip()
            if len(txt) >= 50:
                lig["transcricao"] = txt
                lig["transcricao_fonte"] = "TXT_EXISTENTE" if "TRANSCRIÇÕES" in str(arq) else "MP3_NOVO"
                encontradas += 1
    LOG.info(f"Transcrições casadas: {encontradas}/{sum(1 for l in ligacoes if l['bem_sucedida'])}")
    return encontradas

# ============================================================================
# 8. PROMPTS
# ============================================================================

def prompt_sucesso(lig: Dict, t: str) -> str:
    return f"""Você é analista de qualidade especializado em equipes comerciais de seguros (Ragaz / Suhai).
Avalie esta ligação de SUCESSO DO CORRETOR com foco em resultado, receita e SPIN.

CONTEXTO: Gestor={lig['gestor']} | Duração={lig['duracao_seg']}s | Tipo={lig['tipo']}

TRANSCRIÇÃO (até {MAX_PALAVRAS_TRANSC} palavras):
{t}

Responda APENAS em JSON válido, sem markdown:
{{
  "efetivo": "SIM | NAO | PARCIAL",
  "confianca": 0.0,
  "tipo_pendencia_detectado": "PENDENCIA_PROPOSTA | ATIVACAO | ONBOARDING | OUTRO",
  "tipo_pendencia_especifico": "Recusa Reavaliavel | Rastreador | Desconto | Vistoria | Outro",
  "padrao_comercial": {{"nivel": "ALTO | MEDIO | BAIXO", "descricao": "..."}},
  "spin_aplicado": {{"situacao":"SIM|NAO|PARCIAL","problema":"SIM|NAO|PARCIAL","implicacao":"SIM|NAO|PARCIAL","necessidade":"SIM|NAO|PARCIAL"}},
  "impacto_receita": {{"resultado": "POSITIVO | NEUTRO | NEGATIVO", "motivo": "..."}},
  "boas_praticas": ["..."],
  "pontos_criticos": ["..."],
  "resumo": "uma frase",
  "prioridade_coaching": "ALTA | MEDIA | BAIXA"
}}"""


def prompt_atendimento(lig: Dict, t: str) -> str:
    return f"""Você é especialista em qualidade de ATENDIMENTO ao corretor em seguros (Ragaz).
Avalie no modelo BINÁRIO de 8 critérios.

CONTEXTO: Agente={lig['gestor']} | Duração={lig['duracao_seg']}s

CRITÉRIOS (peso):
1.1 Pronto atendimento e saudação (10)
2.1 Atenção/Concentração (15)
3.1 Entrega de solução efetiva (20)
4.1 Clareza e segurança (15)
5.1 Transferência de contato (10)
6.1 Tempo de espera/Respostas rápidas (10)
7.1 Vocabulário (10)
8.1 Cordialidade/Empatia (10)

nota_final = soma dos pesos cumpridos / 10 (escala 0..10).

TRANSCRIÇÃO:
{t}

Responda APENAS em JSON válido, sem markdown:
{{
  "nota_final": 0.0,
  "criterios": {{
    "pronto_atendimento_saudacao": {{"status": "CUMPRIU | NAO_CUMPRIU", "peso": 10, "evidencia": "..."}},
    "atencao_concentracao":        {{"status": "CUMPRIU | NAO_CUMPRIU", "peso": 15, "evidencia": "..."}},
    "entrega_solucao_efetiva":     {{"status": "CUMPRIU | NAO_CUMPRIU", "peso": 20, "evidencia": "..."}},
    "clareza_seguranca":           {{"status": "CUMPRIU | NAO_CUMPRIU", "peso": 15, "evidencia": "..."}},
    "transferencia_contato":       {{"status": "CUMPRIU | NAO_CUMPRIU", "peso": 10, "evidencia": "..."}},
    "tempo_espera_resposta":       {{"status": "CUMPRIU | NAO_CUMPRIU", "peso": 10, "evidencia": "..."}},
    "vocabulario":                 {{"status": "CUMPRIU | NAO_CUMPRIU", "peso": 10, "evidencia": "..."}},
    "cordialidade_empatia":        {{"status": "CUMPRIU | NAO_CUMPRIU", "peso": 10, "evidencia": "..."}}
  }},
  "pontos_fortes": ["..."],
  "pontos_desenvolvimento": ["..."],
  "resumo": "uma frase",
  "prioridade_coaching": "ALTA | MEDIA | BAIXA"
}}"""


def prompt_comercial(lig: Dict, t: str) -> str:
    return f"""Você é especialista em performance comercial de seguros (Ragaz — Comercial Cadastro).
Avalie com olhar de CLOSER.

CONTEXTO: Agente={lig['gestor']} | Duração={lig['duracao_seg']}s

TRANSCRIÇÃO:
{t}

Responda APENAS em JSON válido, sem markdown:
{{
  "efetivo": "SIM | NAO | PARCIAL",
  "confianca": 0.0,
  "tipo_contato": "CADASTRO_NOVO | OFERTA_BANDEIRA | REATIVACAO | OUTRO",
  "abordagem_comercial": {{
    "abertura": "FORTE | MEDIA | FRACA",
    "qualificacao_corretor": "SIM | NAO | PARCIAL",
    "oferta_clara": "SIM | NAO",
    "tratamento_objecao": "SIM | NAO | SEM_OBJECAO",
    "fechamento": "SIM | NAO | PARCIAL"
  }},
  "resultado_comercial": {{"status": "CONVERTIDO | EM_ANDAMENTO | NAO_CONVERTIDO", "proximo_passo": "..."}},
  "boas_praticas": ["..."],
  "pontos_criticos": ["..."],
  "resumo": "uma frase",
  "prioridade_coaching": "ALTA | MEDIA | BAIXA"
}}"""


def montar_prompt(lig: Dict) -> str:
    t = truncar(lig["transcricao"] or "")
    if lig["area"] == "SUCESSO_CORRETOR":
        return prompt_sucesso(lig, t)
    if lig["area"] == "ATENDIMENTO":
        return prompt_atendimento(lig, t)
    return prompt_comercial(lig, t)

# ============================================================================
# 9. ANÁLISE CLAUDE
# ============================================================================

class ClaudeAnalyzer:
    def __init__(self) -> None:
        from anthropic import Anthropic
        api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
        if not api_key:
            raise RuntimeError(
                "\n[ERRO] ANTHROPIC_API_KEY não configurada.\n"
                "No PowerShell: $env:ANTHROPIC_API_KEY = 'sk-ant-...'\n"
            )
        self.client    = Anthropic(api_key=api_key)
        self.tokens_in  = 0
        self.tokens_out = 0
        self.erros: List[Dict] = []

    def _json(self, texto: str) -> Optional[Dict]:
        try:
            return json.loads(texto)
        except json.JSONDecodeError:
            m = re.search(r"\{.*\}", texto, re.DOTALL)
            if m:
                try:
                    return json.loads(m.group())
                except Exception:
                    pass
        return None

    def analisar(self, lig: Dict) -> Optional[Dict]:
        if not lig.get("transcricao") or len(lig["transcricao"].strip()) < 50:
            return None
        prompt = montar_prompt(lig)
        for tentativa in range(1, MAX_RETRIES_CLAUDE + 1):
            try:
                resp = self.client.messages.create(
                    model=CLAUDE_MODEL,
                    max_tokens=MAX_TOKENS_ANALISE,
                    messages=[{"role": "user", "content": prompt}],
                )
                analise = self._json(resp.content[0].text)
                if analise is None:
                    raise RuntimeError("JSON inválido na resposta")
                self.tokens_in  += resp.usage.input_tokens
                self.tokens_out += resp.usage.output_tokens
                return analise
            except Exception as e:
                LOG.warning(f"Claude tentativa {tentativa}/{MAX_RETRIES_CLAUDE}: {str(e)[:80]}")
                if tentativa < MAX_RETRIES_CLAUDE:
                    time.sleep(CLAUDE_RETRY_DELAY * tentativa)
                else:
                    self.erros.append({"gestor": lig["gestor"], "motivo": str(e)[:120]})
        return None

    def processar_area(self, ligacoes: List[Dict], area_id: str, dry_run: bool) -> int:
        candidatas = [l for l in ligacoes if l["area"] == area_id and l.get("transcricao")]
        nome = AREAS[area_id]["nome"]
        LOG.info(f"[{nome}] {len(candidatas)} ligações com transcrição para analisar.")

        if dry_run:
            LOG.info(f"[{nome}] dry-run: Claude pulado.")
            return 0

        analisadas = 0
        for lig in tqdm(candidatas, desc=nome[:22], unit="lig"):
            result = self.analisar(lig)
            if result:
                lig["analise"] = result
                analisadas += 1
            time.sleep(CALL_DELAY_SEC)

        LOG.info(f"[{nome}] {analisadas}/{len(candidatas)} analisadas.")
        return analisadas

# ============================================================================
# 10. EXCEL
# ============================================================================

_H_FILL = PatternFill("solid", fgColor="1F4E79")
_H_FONT = Font(color="FFFFFF", bold=True)
_S_FILL = PatternFill("solid", fgColor="D9E1F2")
_S_FONT = Font(bold=True)


def _cab(ws, row=1, ncols=8):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _H_FILL
        cell.font = _H_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _flag(a: Dict) -> Optional[bool]:
    if not a:
        return None
    if "efetivo" in a:
        return str(a["efetivo"]).upper() == "SIM"
    if "nota_final" in a:
        try:
            return float(a["nota_final"]) >= 7.0
        except Exception:
            return None
    return None


def gerar_excel(ligacoes: List[Dict], output_dir: Path, ts: str,
                tokens_in: int, tokens_out: int) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    # Resumo
    ws = wb.create_sheet("Resumo", 0)
    ws["A1"] = "RAGAZ — Monitorias Maio/2026 — V3"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    custo_usd = tokens_in * PRECO_INPUT + tokens_out * PRECO_OUTPUT
    row = 3
    for k, v in [("Gerado em", datetime.now().isoformat()), ("Modelo", CLAUDE_MODEL),
                  ("Tokens entrada", tokens_in), ("Tokens saída", tokens_out),
                  ("Custo USD", round(custo_usd, 6)),
                  ("Custo BRL (est.)", round(custo_usd * USD_TO_BRL, 4))]:
        ws.cell(row=row, column=1).value = k
        ws.cell(row=row, column=2).value = v
        row += 1

    row += 1
    for i, h in enumerate(["Área", "Total XLS", "Com transcrição", "Analisadas", "Efetivas", "Taxa %"], 1):
        ws.cell(row=row, column=i).value = h
    _cab(ws, row=row, ncols=6)
    row += 1

    for area_id, info in AREAS.items():
        ligs_a = [l for l in ligacoes if l["area"] == area_id and l["bem_sucedida"]]
        total  = sum(1 for l in ligacoes if l["area"] == area_id)
        com_t  = sum(1 for l in ligs_a if l.get("transcricao"))
        anal   = [l for l in ligs_a if l.get("analise")]
        ef     = sum(1 for l in anal if _flag(l["analise"]) is True)
        taxa   = round(ef / len(anal) * 100, 1) if anal else 0.0
        ws.append([info["nome"], total, com_t, len(anal), ef, taxa])

    for col_letra, w in zip("ABCDEF", [28, 10, 18, 12, 10, 8]):
        ws.column_dimensions[col_letra].width = w

    # Por Gestor
    ws = wb.create_sheet("Por Gestor", 1)
    hdrs = ["Gestor", "Área", "Supervisor", "Total ligações", "Com transcrição", "Analisadas", "Efetivas", "Taxa %"]
    ws.append(hdrs)
    _cab(ws, ncols=len(hdrs))
    for area_id, info in AREAS.items():
        for ag in info["agentes"]:
            ag_n   = norm(ag)
            ligs_g = [l for l in ligacoes if l["area"] == area_id and norm(l["gestor"]) == ag_n]
            if not ligs_g:
                continue
            com_t  = sum(1 for l in ligs_g if l.get("transcricao"))
            anal   = [l for l in ligs_g if l.get("analise")]
            ef     = sum(1 for l in anal if _flag(l["analise"]) is True)
            taxa   = round(ef / len(anal) * 100, 1) if anal else 0.0
            ws.append([ag, info["nome"], info["supervisor"],
                       len(ligs_g), com_t, len(anal), ef, taxa])
    for i in range(1, len(hdrs) + 1):
        ws.column_dimensions[chr(64 + i)].width = 20

    # Por Ligação
    ws = wb.create_sheet("Por Ligação", 2)
    hdrs = ["Área", "Gestor", "Data", "Hora", "Duração(s)", "Telefone",
            "Tipo", "Status", "CRM", "Fonte transcrição",
            "Efetivo/Nota", "Tipo identificado", "Resumo", "Prioridade coaching"]
    ws.append(hdrs)
    _cab(ws, ncols=len(hdrs))
    for l in ligacoes:
        if not l.get("analise"):
            continue
        a = l["analise"]
        efetivo = a.get("efetivo") or str(a.get("nota_final", ""))
        tipo    = (a.get("tipo_pendencia_especifico") or a.get("tipo_contato")
                   or a.get("tipo_pendencia_detectado") or "")
        ws.append([
            AREAS[l["area"]]["nome"], l["gestor"], l["data"], l["hora"],
            l["duracao_seg"], l["telefone"], l["tipo"], l["status"], l["crm"],
            l.get("transcricao_fonte", ""),
            efetivo, tipo, a.get("resumo", ""), a.get("prioridade_coaching", ""),
        ])
    for i, w in enumerate([22,20,12,10,10,18,10,16,12,16,14,20,40,14], 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Por Dia — visão diária por área
    ws = wb.create_sheet("Por Dia", 3)
    hdrs = ["Data", "Área", "Supervisor", "Bem-sucedidas", "Com transcrição",
            "Analisadas", "Efetivas", "Não efetivas", "Taxa %"]
    ws.append(hdrs)
    _cab(ws, ncols=len(hdrs))

    # coleta todas as datas presentes
    datas = sorted({l["data"] for l in ligacoes if l["data"] and l["bem_sucedida"]})
    for data in datas:
        for area_id, info in AREAS.items():
            ligs_d = [l for l in ligacoes
                      if l["data"] == data and l["area"] == area_id and l["bem_sucedida"]]
            if not ligs_d:
                continue
            com_t = sum(1 for l in ligs_d if l.get("transcricao"))
            anal  = [l for l in ligs_d if l.get("analise")]
            ef    = sum(1 for l in anal if _flag(l["analise"]) is True)
            nef   = sum(1 for l in anal if _flag(l["analise"]) is False)
            taxa  = round(ef / len(anal) * 100, 1) if anal else 0.0
            ws.append([data, info["nome"], info["supervisor"],
                       len(ligs_d), com_t, len(anal), ef, nef, taxa])

    for i, w in enumerate([14, 24, 18, 14, 16, 12, 10, 12, 8], 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Padrões & Insights
    ws = wb.create_sheet("Padrões & Insights", 4)
    analisadas_todas = [l for l in ligacoes if l.get("analise")]

    def _sec(titulo: str, row: int) -> int:
        ws.cell(row=row, column=1).value = titulo
        ws.cell(row=row, column=1).font = _S_FONT
        ws.cell(row=row, column=1).fill = _S_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        return row + 1

    row = 1
    ws.cell(row=row, column=1).value = "RAGAZ — Padrões & Insights — Maio/2026"
    ws.cell(row=row, column=1).font = Font(bold=True, size=13)
    ws.merge_cells("A1:E1")
    row += 2

    # ── Prioridade de Coaching ────────────────────────────────────────────
    row = _sec("PRIORIDADE DE COACHING — POR ÁREA", row)
    for i, h in enumerate(["Área", "ALTA", "MÉDIA", "BAIXA", "Total"], 1):
        c = ws.cell(row=row, column=i)
        c.value = h; c.fill = _H_FILL; c.font = _H_FONT
        c.alignment = Alignment(horizontal="center")
    row += 1
    for area_id, info in AREAS.items():
        ligs_a = [l for l in analisadas_todas if l["area"] == area_id]
        alta   = sum(1 for l in ligs_a if str(l["analise"].get("prioridade_coaching","")).upper() == "ALTA")
        media  = sum(1 for l in ligs_a if str(l["analise"].get("prioridade_coaching","")).upper() == "MEDIA")
        baixa  = sum(1 for l in ligs_a if str(l["analise"].get("prioridade_coaching","")).upper() == "BAIXA")
        ws.append([info["nome"], alta, media, baixa, len(ligs_a)])
        row += 1
    row += 1

    # ── Atendimento: critérios ────────────────────────────────────────────
    row = _sec("ATENDIMENTO — CRITÉRIOS (% NÃO CUMPRIU)", row)
    CRITERIOS_AT = [
        ("pronto_atendimento_saudacao", "Pronto atendimento e saudação"),
        ("atencao_concentracao",        "Atenção / Concentração"),
        ("entrega_solucao_efetiva",     "Entrega de solução efetiva"),
        ("clareza_seguranca",           "Clareza e segurança"),
        ("transferencia_contato",       "Transferência de contato"),
        ("tempo_espera_resposta",       "Tempo de espera / Respostas rápidas"),
        ("vocabulario",                 "Vocabulário"),
        ("cordialidade_empatia",        "Cordialidade / Empatia"),
    ]
    for i, h in enumerate(["Critério", "Cumpriu", "Não cumpriu", "Total", "% Falha"], 1):
        c = ws.cell(row=row, column=i)
        c.value = h; c.fill = _H_FILL; c.font = _H_FONT
        c.alignment = Alignment(horizontal="center")
    row += 1
    ligs_at = [l for l in analisadas_todas if l["area"] == "ATENDIMENTO"]
    for chave, label in CRITERIOS_AT:
        cumpriu = nao = 0
        for l in ligs_at:
            crit = l["analise"].get("criterios", {}).get(chave, {})
            st   = str(crit.get("status", "")).upper()
            if st == "CUMPRIU":
                cumpriu += 1
            elif st == "NAO_CUMPRIU":
                nao += 1
        total = cumpriu + nao
        pct   = round(nao / total * 100, 1) if total else 0.0
        ws.append([label, cumpriu, nao, total, pct])
        row += 1
    row += 1

    # ── Sucesso: tipos de pendência ───────────────────────────────────────
    row = _sec("SUCESSO DO CORRETOR — TIPOS DE PENDÊNCIA MAIS FREQUENTES", row)
    for i, h in enumerate(["Tipo de Pendência", "Qtd", "% do total"], 1):
        c = ws.cell(row=row, column=i)
        c.value = h; c.fill = _H_FILL; c.font = _H_FONT
        c.alignment = Alignment(horizontal="center")
    row += 1
    ligs_sc = [l for l in analisadas_todas if l["area"] == "SUCESSO_CORRETOR"]
    from collections import Counter
    tipos_pend = Counter(
        str(l["analise"].get("tipo_pendencia_especifico") or
            l["analise"].get("tipo_pendencia_detectado") or "Não identificado")
        for l in ligs_sc
    )
    total_sc = len(ligs_sc) or 1
    for tipo, qtd in tipos_pend.most_common():
        ws.append([tipo, qtd, round(qtd / total_sc * 100, 1)])
        row += 1
    row += 1

    # ── Sucesso: SPIN ─────────────────────────────────────────────────────
    row = _sec("SUCESSO DO CORRETOR — APLICAÇÃO DO SPIN", row)
    for i, h in enumerate(["Etapa SPIN", "SIM", "PARCIAL", "NÃO", "% SIM"], 1):
        c = ws.cell(row=row, column=i)
        c.value = h; c.fill = _H_FILL; c.font = _H_FONT
        c.alignment = Alignment(horizontal="center")
    row += 1
    for etapa in ["situacao", "problema", "implicacao", "necessidade"]:
        sim = parc = nao = 0
        for l in ligs_sc:
            v = str(l["analise"].get("spin_aplicado", {}).get(etapa, "")).upper()
            if v == "SIM":     sim  += 1
            elif v == "PARCIAL": parc += 1
            else:              nao  += 1
        total_e = sim + parc + nao or 1
        ws.append([etapa.capitalize(), sim, parc, nao, round(sim / total_e * 100, 1)])
        row += 1
    row += 1

    # ── Comercial: funil ──────────────────────────────────────────────────
    row = _sec("COMERCIAL CADASTRO — FUNIL DE ABORDAGEM", row)
    for i, h in enumerate(["Etapa", "SIM / FORTE", "PARCIAL / MÉDIA", "NÃO / FRACA", "% OK"], 1):
        c = ws.cell(row=row, column=i)
        c.value = h; c.fill = _H_FILL; c.font = _H_FONT
        c.alignment = Alignment(horizontal="center")
    row += 1
    ligs_cc = [l for l in analisadas_todas if l["area"] == "COMERCIAL_CADASTRO"]
    FUNIL = [
        ("abertura",              "Abertura",              ["FORTE"],   ["MEDIA"],   ["FRACA"]),
        ("qualificacao_corretor", "Qualificação corretor", ["SIM"],     ["PARCIAL"], ["NAO"]),
        ("oferta_clara",          "Oferta clara",          ["SIM"],     [],          ["NAO"]),
        ("tratamento_objecao",    "Tratamento de objeção", ["SIM"],     ["SEM_OBJECAO"], ["NAO"]),
        ("fechamento",            "Fechamento",            ["SIM"],     ["PARCIAL"], ["NAO"]),
    ]
    for chave, label, pos, parc_v, neg in FUNIL:
        ok = pa = no = 0
        for l in ligs_cc:
            v = str(l["analise"].get("abordagem_comercial", {}).get(chave, "")).upper()
            if v in pos:    ok += 1
            elif v in parc_v: pa += 1
            else:           no += 1
        total_f = ok + pa + no or 1
        ws.append([label, ok, pa, no, round(ok / total_f * 100, 1)])
        row += 1

    for col_letra, w in zip("ABCDE", [36, 12, 14, 12, 10]):
        ws.column_dimensions[col_letra].width = w

    # Sem transcrição
    ws = wb.create_sheet("Sem Transcrição", 5)
    hdrs = ["Área", "Gestor", "Data", "Hora", "Telefone", "Status"]
    ws.append(hdrs)
    _cab(ws, ncols=len(hdrs))
    for l in ligacoes:
        if l["bem_sucedida"] and not l.get("transcricao"):
            ws.append([AREAS[l["area"]]["nome"], l["gestor"],
                       l["data"], l["hora"], l["telefone"], l["status"]])
    for i in range(1, len(hdrs) + 1):
        ws.column_dimensions[chr(64 + i)].width = 20

    destino = output_dir / f"monitorias_maio_{ts}.xlsx"
    wb.save(destino)
    return destino

# ============================================================================
# 11. MAIN
# ============================================================================

def main() -> int:
    parser = argparse.ArgumentParser(description="RAGAZ Monitorias V3")
    parser.add_argument("--base",  default=str(BASE_DIR))
    parser.add_argument("--area",
                        choices=["TODAS","SUCESSO_CORRETOR","ATENDIMENTO","COMERCIAL_CADASTRO"],
                        default="TODAS")
    parser.add_argument("--max",      type=int, default=0)
    parser.add_argument("--data",     default="",
                        help="Filtra só uma data. Formato: YYYY-MM-DD (ex: 2026-05-14)")
    parser.add_argument("--dry-run",  action="store_true")
    parser.add_argument("--so-txts",  action="store_true",
                        help="Pula transcrição de MP3 — usa só TXTs existentes")
    args = parser.parse_args()

    base        = Path(args.base)
    transcr_dir = base / "TRANSCRIÇÕES"
    audios_dir  = base / "AUDIOS"
    output_dir  = base / f"RELATÓRIOS {datetime.now():%d.%m.%y}"
    output_dir.mkdir(parents=True, exist_ok=True)
    _add_file_log(output_dir)

    LOG.info("=" * 70)
    LOG.info(f"RAGAZ MONITORIAS V3 — {datetime.now():%Y-%m-%d %H:%M:%S}")
    LOG.info(f"Área: {args.area}  |  dry-run: {args.dry_run}  |  só-txts: {args.so_txts}")
    LOG.info("=" * 70)

    # 1) XLS
    xls_path = encontrar_xls(base)
    if not xls_path:
        LOG.error(f"Nenhum XLS encontrado em {base}")
        return 2
    df       = ler_xls(xls_path)
    ligacoes = processar_xls(df)

    if args.area != "TODAS":
        ligacoes = [l for l in ligacoes if l["area"] == args.area]

    bem_suc = [l for l in ligacoes if l["bem_sucedida"]]
    if args.data:
        bem_suc = [l for l in bem_suc if l["data"] == args.data]
        LOG.info(f"Filtro por data: {args.data} → {len(bem_suc)} ligações")
    if args.max and args.max > 0:
        bem_suc = bem_suc[:args.max]
    LOG.info(f"Bem-sucedidas elegíveis: {len(bem_suc)}")

    # 2) Indexar arquivos
    indice = IndiceArquivos(transcr_dir, audios_dir)

    # 3) Transcrever MP3s pendentes (pulado em dry-run)
    if not args.so_txts and not args.dry_run:
        transcriber = Transcriber(transcr_dir)
        transcriber.processar_pendentes(indice, bem_suc)

    # 4) Casar transcrições
    casar_transcricoes(bem_suc, indice)

    # 5) Análise Claude por área
    analyzer    = ClaudeAnalyzer()
    areas_rodar = [args.area] if args.area != "TODAS" else list(AREAS.keys())
    total_anal  = 0

    for i, area_id in enumerate(areas_rodar):
        n = analyzer.processar_area(bem_suc, area_id, args.dry_run)
        total_anal += n
        if i < len(areas_rodar) - 1 and not args.dry_run:
            LOG.info(f"Pausa {AREA_PAUSE_SEC}s entre áreas (reset rate limit)...")
            for restante in range(AREA_PAUSE_SEC, 0, -10):
                LOG.info(f"  ... {restante}s")
                time.sleep(10)

    # 6) Relatórios
    ts        = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = gerar_excel(bem_suc, output_dir, ts, analyzer.tokens_in, analyzer.tokens_out)

    custo_usd = analyzer.tokens_in * PRECO_INPUT + analyzer.tokens_out * PRECO_OUTPUT
    json_data = {
        "gerado_em":   datetime.now().isoformat(),
        "modelo":      CLAUDE_MODEL,
        "tokens_in":   analyzer.tokens_in,
        "tokens_out":  analyzer.tokens_out,
        "custo_usd":   round(custo_usd, 6),
        "custo_brl":   round(custo_usd * USD_TO_BRL, 4),
        "erros_claude":analyzer.erros,
        "areas": {
            area_id: {
                "nome":              info["nome"],
                "total_xls":         sum(1 for l in ligacoes if l["area"] == area_id),
                "bem_sucedidas":     sum(1 for l in bem_suc if l["area"] == area_id),
                "com_transcricao":   sum(1 for l in bem_suc if l["area"] == area_id and l.get("transcricao")),
                "analisadas":        sum(1 for l in bem_suc if l["area"] == area_id and l.get("analise")),
                "efetivas":          sum(1 for l in bem_suc if l["area"] == area_id
                                         and _flag(l.get("analise")) is True),
            }
            for area_id, info in AREAS.items()
        },
    }
    json_path = output_dir / f"monitorias_maio_{ts}.json"
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(json_data, f, ensure_ascii=False, indent=2, default=str)

    # 7) Resumo console
    print("\n" + "=" * 70)
    print("  RESUMO — RAGAZ MONITORIAS V3")
    print("=" * 70)
    print(f"  XLS               : {xls_path.name}")
    print(f"  Bem-sucedidas     : {len(bem_suc)}")
    print(f"  Total analisadas  : {total_anal}")
    print(f"  Tokens (in/out)   : {analyzer.tokens_in} / {analyzer.tokens_out}")
    print(f"  Custo             : USD {round(custo_usd,4)}  (≈ BRL {round(custo_usd*USD_TO_BRL,2)})")
    print(f"  Erros Claude      : {len(analyzer.erros)}")
    print()
    print("  POR ÁREA:")
    for area_id, d in json_data["areas"].items():
        ef   = d["efetivas"]
        anal = d["analisadas"]
        taxa = round(ef / anal * 100, 1) if anal else 0
        print(f"   - {d['nome']:<24}  transcr={d['com_transcricao']:>4}  "
              f"anal={anal:>4}  efet={ef:>4}  {taxa:>5}%")
    print()
    print(f"  Arquivos: {xlsx_path.name}")
    print(f"            {json_path.name}")
    print("=" * 70)
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        LOG.warning("Interrompido (Ctrl+C).")
        sys.exit(130)
    except Exception as e:
        LOG.error(f"ERRO FATAL: {e}")
        traceback.print_exc()
        sys.exit(1)
